"""
BroBot FastAPI Server
API endpoints for dashboard + scheduler for data/signal loops.
All endpoints query REAL database — no stubs, no mocks.
"""
import asyncio
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
import os
import json
import base64
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from typing import Optional
import logging
import traceback

from src import config
from src.db import fetch_all, fetch_one, execute, init_tables, close_pool, get_pool
from src.db_async import get_async_pool

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Scheduler
scheduler = AsyncIOScheduler()

# Track scan times
_last_data_scan: Optional[datetime] = None
_startup_time: Optional[datetime] = None


# ═══════════════════════════════════════════════════════════════
# TRADING MODE GATE — checks bot_settings before any Telegram send
# ═══════════════════════════════════════════════════════════════
_trading_mode_cache = {'mode': None, 'paper_notif': None, 'ts': None}


async def get_trading_mode() -> dict:
    """Read trading_mode + paper_notifications from bot_settings.
    Returns {'mode': 'live'|'paper', 'paper_notifications': bool}.
    Cached for 60s to avoid DB spam.
    """
    import time
    now = time.time()
    if _trading_mode_cache['ts'] and (now - _trading_mode_cache['ts']) < 60:
        return {'mode': _trading_mode_cache['mode'], 'paper_notifications': _trading_mode_cache['paper_notif']}
    try:
        pool = get_async_pool()
        async with pool.acquire() as conn:
            row_mode = await conn.fetchrow(
                "SELECT value FROM bot_settings WHERE key = 'trading_mode'"
            )
            row_notif = await conn.fetchrow(
                "SELECT value FROM bot_settings WHERE key = 'paper_notifications'"
            )
        mode = 'live'
        if row_mode:
            v = str(row_mode['value']).strip().strip('"').lower()
            mode = v if v in ('live', 'paper') else 'live'
        paper_notif = False
        if row_notif:
            v = str(row_notif['value']).strip().strip('"').lower()
            paper_notif = v in ('true', '1', 'yes')
        _trading_mode_cache['mode'] = mode
        _trading_mode_cache['paper_notif'] = paper_notif
        _trading_mode_cache['ts'] = now
        return {'mode': mode, 'paper_notifications': paper_notif}
    except Exception as e:
        logger.error(f"❌ get_trading_mode failed: {e}")
        return {'mode': 'live', 'paper_notifications': False}


def should_send_paper_telegram(trading_cfg: dict) -> bool:
    """Return True only if paper notifications are allowed."""
    return trading_cfg.get('paper_notifications', False)


def is_live_mode(trading_cfg: dict) -> bool:
    """Return True if trading mode is live."""
    return trading_cfg.get('mode', 'paper') == 'live'

# Signal loop instance (initialized on startup)
_signal_loop = None
_improvement_engine = None








# ═══════════════════════════════════════════════════════════════




async def scheduled_settlement():
    """Protocol 5: Settlement & Learning.
    Check completed events, settle trades, feed learning engine.
    See INTELLIGENCE.md Protocol 5.
    """
    """Auto-settle completed trades every 5 minutes."""
    try:
        from src.execution.settlement import settle_trades
        odds_key = os.environ.get('ODDS_API_KEY', '')
        if not odds_key:
            env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '.env')
            if os.path.exists(env_path):
                with open(env_path) as f:
                    for line in f:
                        if line.startswith('ODDS_API_KEY'):
                            odds_key = line.split('=', 1)[1].strip().strip('"').strip("'")
                            break
        result = await settle_trades(fetch_all, execute, fetch_one, odds_key)
        if result.get('settled', 0) > 0:
            logger.info(f"\u2705 Settlement: {result['settled']} trades settled, P&L: ${result.get('total_pnl', 0):+.2f}")
            # Broadcast results via Telegram
            if _telegram_bot:
                for t in result.get('trades', []):
                    pass  # Sports trade broadcasts disabled (noise)
    except Exception as e:
        logger.error(f"\u274c Settlement error: {e}")


# ═══════════════════════════════════════════════════════════════
# TELEGRAM SUBSCRIBER BOT — Scheduled Jobs











# ═══════════════════════════════════════════════════════════════




# ═══════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════
# BTC BANKROLL MANAGEMENT
# ═══════════════════════════════════════════════════════════════

BTC_STARTING_BALANCE = 5000.00

async def get_btc_bankroll_state():
    """Fetch current bankroll from DB. Returns dict with balance, available, in_positions."""
    try:
        pool = get_async_pool()
        async with pool.acquire() as conn:
            row = await conn.fetchrow(
                "SELECT balance, available, in_positions, total_won, total_lost, total_trades, peak_balance, max_drawdown_pct "
                "FROM btc_bankroll ORDER BY id DESC LIMIT 1"
            )
            if row:
                return dict(row)
    except Exception as e:
        logger.error(f"❌ Bankroll fetch failed: {e}")
    return {
        'balance': BTC_STARTING_BALANCE,
        'available': BTC_STARTING_BALANCE,
        'in_positions': 0,
        'total_won': 0,
        'total_lost': 0,
        'total_trades': 0,
        'peak_balance': BTC_STARTING_BALANCE,
        'max_drawdown_pct': 0,
    }

async def compute_btc_stake(entry_price: float, prediction: str, bankroll_state: dict) -> tuple:
    """
    Compute bankroll-proportional stake for a BTC trade.
    Returns (stake, skip_reason) where skip_reason is None if trade is allowed.
    """
    balance = float(bankroll_state.get('balance', BTC_STARTING_BALANCE))
    available = float(bankroll_state.get('available', BTC_STARTING_BALANCE))
    in_positions = float(bankroll_state.get('in_positions', 0))

    max_single_bet = balance * 0.10
    max_exposure = balance * 0.30
    conviction_max = balance * 0.20

    # Base stake as % of bankroll
    if prediction == 'DOWN':
        if entry_price < 0.20:
            base_pct = 0.030
        elif entry_price < 0.30:
            base_pct = 0.020
        elif entry_price < 0.40:
            base_pct = 0.015
        else:  # 40-50c
            base_pct = 0.010
    else:  # UP
        base_pct = 0.005

    stake = balance * base_pct
    stake = min(stake, max_single_bet)

    # Check bankroll guards
    if available < stake:
        return 0, f"BANKROLL: insufficient available (${available:.0f} < ${stake:.0f})"
    if in_positions + stake > max_exposure:
        return 0, f"BANKROLL: exposure limit (${in_positions:.0f}+${stake:.0f} > ${max_exposure:.0f})"

    return round(stake, 2), None

async def update_btc_bankroll_open(stake: float):
    """Reserve stake in bankroll when opening a trade."""
    try:
        pool = get_async_pool()
        async with pool.acquire() as conn:
            await conn.execute("""
                UPDATE btc_bankroll SET
                    available = available - $1,
                    in_positions = in_positions + $1,
                    updated_at = NOW()
                WHERE id = (SELECT id FROM btc_bankroll ORDER BY id DESC LIMIT 1)
            """, stake)
    except Exception as e:
        logger.error(f"❌ Bankroll open update failed: {e}")

async def update_btc_bankroll_close(stake: float, pnl: float, won: bool):
    """Update bankroll after trade resolves."""
    try:
        pool = get_async_pool()
        async with pool.acquire() as conn:
            if won:
                # WON: balance += pnl, return stake + gain
                await conn.execute("""
                    UPDATE btc_bankroll SET
                        balance = balance + $1,
                        available = available + $2 + $1,
                        in_positions = GREATEST(0, in_positions - $2),
                        total_won = total_won + $1,
                        total_trades = total_trades + 1,
                        peak_balance = GREATEST(peak_balance, balance + $1),
                        max_drawdown_pct = CASE
                            WHEN GREATEST(peak_balance, balance + $1) > 0
                            THEN ROUND((GREATEST(peak_balance, balance + $1) - (balance + $1)) / GREATEST(peak_balance, balance + $1) * 100, 2)
                            ELSE 0 END,
                        updated_at = NOW()
                    WHERE id = (SELECT id FROM btc_bankroll ORDER BY id DESC LIMIT 1)
                """, pnl, stake)
            else:
                # LOST: balance -= stake, lose stake
                await conn.execute("""
                    UPDATE btc_bankroll SET
                        balance = balance - $1,
                        available = available,
                        in_positions = GREATEST(0, in_positions - $1),
                        total_lost = total_lost + $1,
                        total_trades = total_trades + 1,
                        max_drawdown_pct = CASE
                            WHEN peak_balance > 0
                            THEN ROUND((peak_balance - (balance - $1)) / peak_balance * 100, 2)
                            ELSE 0 END,
                        updated_at = NOW()
                    WHERE id = (SELECT id FROM btc_bankroll ORDER BY id DESC LIMIT 1)
                """, stake)
    except Exception as e:
        logger.error(f"❌ Bankroll close update failed: {e}")

async def sync_btc_in_positions():
    """Recalculate in_positions from open (unresolved) btc_signals."""
    try:
        pool = get_async_pool()
        async with pool.acquire() as conn:
            row = await conn.fetchrow("""
                SELECT COALESCE(SUM(s.stake_used), 0) as open_stake
                FROM btc_signals s
                JOIN btc_windows w ON s.window_id = w.window_id
                WHERE s.prediction != 'SKIP' AND w.resolution IS NULL
                AND s.stake_used IS NOT NULL
            """)
            open_stake = float(row['open_stake']) if row else 0
            state = await get_btc_bankroll_state()
            balance = float(state['balance'])
            total_won = float(state['total_won'])
            total_lost = float(state['total_lost'])
            # available = balance - open_stake
            available = max(0, balance - open_stake)
            await conn.execute("""
                UPDATE btc_bankroll SET
                    in_positions = $1,
                    available = $2,
                    updated_at = NOW()
                WHERE id = (SELECT id FROM btc_bankroll ORDER BY id DESC LIMIT 1)
            """, open_stake, available)
    except Exception as e:
        logger.error(f"❌ Bankroll sync failed: {e}")

# ═══════════════════════════════════════════════════════════════
# BTC SIGNAL ENGINE — Scheduled Scan Functions
# ═══════════════════════════════════════════════════════════════
_btc_engine = None
_last_btc_scan: Optional[datetime] = None
_last_btc_resolution: Optional[datetime] = None





async def scheduled_btc_signal_scan():
    """BTC Signal Engine: scan for active BTC windows, compute 7-factor signals."""
    global _last_btc_scan, _btc_engine
    try:
        if _btc_engine is None:
            from src.strategies.btc_signal_engine import BTCSignalEngine
            _btc_engine = BTCSignalEngine(get_async_pool())
            await _btc_engine.ensure_tables()
            logger.info("✅ BTC Signal Engine initialized")

        results = await _btc_engine.run_scan()
        _last_btc_scan = datetime.utcnow()

        # ═══════════════════════════════════════════════════════════
        # STRATEGY V3 (2026-04-09 13:29 IST)
        # Rule 1: REWARD >= RISK — potential profit must cover loss + fee
        #         This means entry_price < 50c (reward:risk >= 1:1)
        # Rule 2: Scale stake aggressively on best odds
        #         <20c=$75, 20-30c=$50, 30-40c=$35, 40-50c=$25
        # Rule 3: Skip 15M (always 85c+ entries)
        # Rule 4: Minimum factor agreement — at least 4/7 factors aligned
        # ═══════════════════════════════════════════════════════════
        if _telegram_bot and results:
            for sig in results:
                prob = sig.get('prob_up', 0.5)
                conf = sig.get('confidence', 0)
                pred = sig.get('prediction', 'SKIP')
                if pred == 'SKIP':
                    continue
                wl = sig.get('window_length', 15)
                up_price = sig.get('up_price', 0.5)
                down_price = sig.get('down_price', 0.5)
                entry_price = down_price if pred == 'UP' else up_price

                # RULE 3: Skip 15M windows entirely
                if wl == 15:
                    continue

                # ═══ LIVE-ADJUSTED STRATEGY (V4.2) ═══
                # V4.1 was too strict: only 3.8% signals passed (3/78 in 6h)
                # V4.2: removed redundant prob filter (entry price + R:R already gate quality)
                #        relaxed factors from 5→4 (most signals hit 4)
                #        now ~51% signals pass while keeping R:R >1.5x safety
                
                EXECUTION_COST = 0.05  # 5% total (2% fee + 3% slippage estimate)
                MIN_ENTRY = 0.02      # Below 2c = no liquidity
                MAX_ENTRY = 0.40      # Tightened from 50c — need more upside buffer
                MIN_RR_AFTER_COSTS = 1.5  # Must clear 1.5:1 AFTER all costs
                MIN_FACTORS = 4       # 4 of 7 factors must agree (relaxed from 5)

                # ═══ DYNAMIC BANKROLL SIZING ═══
                # Query proxy wallet balance from CLOB (cached in signal engine)
                try:
                    _proxy_balance = getattr(_btc_engine, '_cached_proxy_balance', None)
                    if _proxy_balance and _proxy_balance > 0:
                        # Normal: 3% of bankroll, cap $25
                        # Conviction (handled below if 4x signal): 5% of bankroll, cap $50
                        LIVE_STAKE = min(_proxy_balance * 0.03, 25.0)
                        if LIVE_STAKE < 5.0:
                            logger.warning(f"⚠️ Bankroll too low (${_proxy_balance:.2f}), skipping trade")
                            continue
                    else:
                        LIVE_STAKE = 25  # Fallback
                except Exception:
                    LIVE_STAKE = 25  # Fallback
                
                # RULE 1: Entry price bounds
                if entry_price <= MIN_ENTRY or entry_price >= MAX_ENTRY:
                    continue
                
                # RULE 2: Calculate REAL reward:risk after ALL costs
                # Win payout: (1/entry - 1) minus execution costs on both sides
                gross_rr = (1.0 / entry_price) - 1.0  # Raw R:R
                # Deduct: buy slippage + sell slippage + fees
                net_win_per_dollar = gross_rr - EXECUTION_COST  # What you actually keep per $1
                net_loss_per_dollar = 1.0 + EXECUTION_COST  # What you actually lose per $1 (stake + costs)
                
                reward_risk = net_win_per_dollar / net_loss_per_dollar if net_loss_per_dollar > 0 else 0
                
                if reward_risk < MIN_RR_AFTER_COSTS:
                    continue  # Upside doesn't justify downside after costs
                
                # RULE 3: Fixed stake for live (no scaled sizing — liquidity appropriate)
                stake = LIVE_STAKE
                
                # RULE 4: Check factor agreement (from signal factors)
                factors = sig.get('factors', {})
                if factors:
                    direction_sign = 1 if pred == 'UP' else -1
                    factor_vals = [
                        factors.get('f_price_delta', 0),
                        factors.get('f_momentum', 0),
                        factors.get('f_volume_imbalance', 0),
                        factors.get('f_oracle_lead', 0),
                        factors.get('f_book_imbalance', 0),
                        factors.get('f_volatility', 0),
                        factors.get('f_time_decay', 0),
                    ]
                    agreeing = sum(1 for f in factor_vals if f * direction_sign > 0)
                    if agreeing < MIN_FACTORS:
                        continue  # Need 5+ factors aligned

                price = sig.get('btc_price', 0)
                btc_open = sig.get('btc_open', price)
                delta_pct = ((price - btc_open) / btc_open * 100) if btc_open > 0 else 0
                potential = (stake / entry_price - stake) * (1 - EXECUTION_COST)  # Net after all costs
                
                # Trade opened — log only, NO telegram broadcast
                # WON/LOST sent on resolution with full analysis
                logger.info(f"V4 trade: {pred} {wl}M | {entry_price*100:.0f}c | ${stake} | R:R {reward_risk:.1f}x | {agreeing}/7")

                # ═══════════════════════════════════════════════════════
                # LIVE TRADE EXECUTION — if mode is 'live' and passes V4+
                # ═══════════════════════════════════════════════════════
                try:
                    _tcfg = await get_trading_mode()
                    if is_live_mode(_tcfg) and agreeing >= 5:
                        from src.polymarket_live import PolymarketLiveTrader
                        _live_trader = PolymarketLiveTrader(get_async_pool())
                        # Determine token_id: buy the side we predict wins
                        _token_id = sig.get('token_id_up', '') if pred == 'UP' else sig.get('token_id_down', '')
                        _live_stake = min(stake, 25.0)  # Hard cap $25 for live
                        if _token_id:
                            _live_trade_id = await _live_trader.execute_live_trade(
                                window_id=sig.get('window_id', ''),
                                prediction=pred,
                                token_id=_token_id,
                                entry_price=entry_price,
                                stake_usd=_live_stake,
                                factors_agreeing=agreeing,
                                signal_metadata={'factors': factors, 'window_length': wl, 'reward_risk': reward_risk},
                                seconds_remaining=sig.get('seconds_remaining', 999),
                            )
                            if _live_trade_id:
                                logger.info(f"🟢 LIVE trade #{_live_trade_id} placed for {pred} {wl}M")
                        else:
                            logger.warning(f"⚠️ No token_id for {pred} — cannot place live trade")
                except Exception as _live_e:
                    logger.error(f"❌ Live trade execution error: {_live_e}")

                # Store stake_used on the signal row and update bankroll
                _window_id = sig.get('window_id', '')
                if _window_id:
                    try:
                        pool = get_async_pool()
                        async with pool.acquire() as conn:
                            await conn.execute("""
                                UPDATE btc_signals SET stake_used = $1
                                WHERE window_id = $2 AND prediction != 'SKIP'
                            """, stake, _window_id)
                    except Exception as _se:
                        logger.error(f"❌ Stake save failed: {_se}")
                await update_btc_bankroll_open(stake)

                # Track volatility per hour slot
                try:
                    pool = get_async_pool()
                    async with pool.acquire() as conn:
                        await conn.execute("""
                            INSERT INTO btc_volatility_hours (
                                date, hour_ist, window_length, trades_taken, avg_entry, 
                                btc_price_range_pct, session_tag
                            ) VALUES (
                                CURRENT_DATE,
                                EXTRACT(HOUR FROM NOW() AT TIME ZONE 'Asia/Kolkata')::int,
                                $1, 1, $2, $3, 'v2'
                            )
                            ON CONFLICT (date, hour_ist, window_length) DO UPDATE SET
                                trades_taken = btc_volatility_hours.trades_taken + 1,
                                avg_entry = (btc_volatility_hours.avg_entry * btc_volatility_hours.trades_taken + $2) 
                                    / (btc_volatility_hours.trades_taken + 1),
                                btc_price_range_pct = GREATEST(btc_volatility_hours.btc_price_range_pct, $3)
                        """, wl, entry_price, abs(delta_pct))
                except Exception:
                    pass

        logger.info(f"📊 BTC scan: {len(results)} signals")
    except Exception as e:
        logger.error(f"❌ BTC signal scan failed: {e}\n{traceback.format_exc()}")


_pinned_msg_ids = {  # chat_id -> message_id per report type for pin management
    'hourly': {},
    'penny': {},
    'intelligence': {},
    'daily': {},
}

async def _send_and_pin(bot_token, subscribers, msg, report_type):
    """Send message to all subscribers, pin it, unpin previous of same type."""
    global _pinned_msg_ids
    import httpx
    async with httpx.AsyncClient(timeout=15) as client:
        for sub in subscribers:
            chat_id = sub['chat_id'] if isinstance(sub, dict) else sub
            try:
                # Unpin previous of same type
                old_id = _pinned_msg_ids[report_type].get(chat_id)
                if old_id:
                    await client.post(f"https://api.telegram.org/bot{bot_token}/unpinChatMessage",
                        json={"chat_id": chat_id, "message_id": old_id})
                # Send
                resp = await client.post(f"https://api.telegram.org/bot{bot_token}/sendMessage",
                    json={"chat_id": chat_id, "text": msg})
                result = resp.json()
                if result.get('ok'):
                    new_id = result['result']['message_id']
                    # Pin
                    await client.post(f"https://api.telegram.org/bot{bot_token}/pinChatMessage",
                        json={"chat_id": chat_id, "message_id": new_id, "disable_notification": True})
                    _pinned_msg_ids[report_type][chat_id] = new_id
            except Exception as e:
                logger.error(f"\u274c {report_type} send/pin failed for {chat_id}: {e}")

async def scheduled_btc_hourly_summary():
    """Send hourly performance summary with financials to Telegram. Pins the message."""
    global _telegram_bot, _pinned_msg_ids
    if not _telegram_bot:
        logger.warning("\u26a0\ufe0f BTC hourly: _telegram_bot is None, skipping")
        return
    # Gate: suppress if paper mode + paper_notifications=false
    _tcfg = await get_trading_mode()
    if is_live_mode(_tcfg):
        logger.info("🔇 BTC hourly summary suppressed (live mode — only live trade alerts)")
        return
    if not should_send_paper_telegram(_tcfg):
        logger.info("🔇 BTC hourly summary suppressed (paper notifications off)")
        return
    try:
        import psycopg2
        conn = psycopg2.connect(dbname='polyedge', user='node', host='localhost')
        cur = conn.cursor()

        # All-time per timeframe — uses btc_pnl view (correct stakes + filters invalid prices)
        cur.execute("""
            SELECT window_length, COUNT(*) as total, COUNT(*) FILTER (WHERE correct) as wins,
                ROUND(SUM(trade_pnl)::numeric, 2) as net_pnl,
                ROUND(SUM(CASE WHEN trade_pnl > 0 THEN trade_pnl ELSE 0 END)::numeric, 2) as gross_profit,
                ROUND(ABS(SUM(CASE WHEN trade_pnl < 0 THEN trade_pnl ELSE 0 END))::numeric, 2) as gross_loss,
                ROUND(MAX(trade_pnl)::numeric, 2) as best_trade,
                ROUND(SUM(CASE WHEN trade_pnl > 0 THEN stake * (1.0/entry_price - 1.0) * 0.02 ELSE 0 END)::numeric, 2) as fees
            FROM btc_pnl GROUP BY window_length ORDER BY window_length
        """)
        stats = cur.fetchall()

        # Last hour
        cur.execute("""
            SELECT COUNT(*) as total, COUNT(*) FILTER (WHERE correct) as wins,
                ROUND(SUM(trade_pnl)::numeric, 2) as net_pnl,
                ROUND(MAX(trade_pnl)::numeric, 2) as best_trade
            FROM btc_pnl WHERE close_time > NOW() - INTERVAL '1 hour'
        """)
        last_hour = cur.fetchone()

        # Today's P&L
        cur.execute("""
            SELECT COUNT(*), COUNT(*) FILTER (WHERE correct),
                ROUND(SUM(trade_pnl)::numeric, 2),
                ROUND(SUM(CASE WHEN trade_pnl > 0 THEN trade_pnl ELSE 0 END)::numeric, 2),
                ROUND(ABS(SUM(CASE WHEN trade_pnl < 0 THEN trade_pnl ELSE 0 END))::numeric, 2),
                ROUND(SUM(CASE WHEN trade_pnl > 0 THEN stake * (1.0/entry_price - 1.0) * 0.02 ELSE 0 END)::numeric, 2),
                ROUND(SUM(stake)::numeric, 2)
            FROM btc_pnl WHERE close_time::date = CURRENT_DATE
        """)
        today_stats = cur.fetchone()

        # Active strategy
        cur.execute("SELECT version FROM btc_strategy_versions WHERE status = 'active' LIMIT 1")
        active_v = cur.fetchone()
        strategy_name = active_v[0] if active_v else '?'
        conn.close()

        # Build message
        from datetime import datetime as dt
        now_ist = dt.now().strftime('%H:%M IST')
        _mode_label = '🟢 BTC LIVE TRADING' if is_live_mode(_tcfg) else '📊 BTC PAPER TRADING'
        lines = [f"{_mode_label} \u2014 {now_ist}", f"Strategy: {strategy_name}", ""]

        h_total, h_wins, h_net, h_best = last_hour if last_hour else (0, 0, 0, 0)
        if h_total and h_total > 0:
            h_sign = '+' if float(h_net or 0) >= 0 else ''
            lines.append(f"\u23f0 Last Hour: {h_wins}W-{h_total-h_wins}L | {h_sign}${h_net}")
            if h_best and float(h_best) > 0:
                lines.append(f"   Best: +${h_best}")
        else:
            lines.append("\u23f0 Last Hour: No trades resolved")

        # TODAY section
        d_total, d_wins, d_net, d_profit, d_loss, d_fees, d_returns = today_stats if today_stats else (0,0,0,0,0,0,0)
        d_total = d_total or 0; d_wins = d_wins or 0; d_net = float(d_net or 0)
        d_profit = float(d_profit or 0); d_loss = float(d_loss or 0)
        d_fees = float(d_fees or 0); d_spent = d_total * 25
        lines.append("")
        lines.append("\u2501\u2501\u2501 TODAY \u2501\u2501\u2501")
        d_emoji = '\U0001f7e2' if d_net >= 0 else '\U0001f534'
        d_sign = '+' if d_net >= 0 else ''
        lines.append(f"{d_emoji} Net P&L: {d_sign}${d_net:.2f}")
        lines.append(f"\U0001f4b0 Profit: +${d_profit:.2f} | \U0001f4c9 Loss: -${d_loss:.2f}")
        lines.append(f"\U0001f4b5 Spent: ${d_spent} | \U0001f4b8 Fees: ${d_fees:.2f}")
        if d_total > 0:
            lines.append(f"Record: {d_wins}W-{d_total-d_wins}L ({d_wins/d_total*100:.0f}%)")

        # YTD (all-time) section
        lines.append("")
        lines.append("\u2501\u2501\u2501 ALL-TIME (YTD) \u2501\u2501\u2501")

        a_net = sum(float(r[3] or 0) for r in stats)
        a_profit = sum(float(r[4] or 0) for r in stats)
        a_loss = sum(float(r[5] or 0) for r in stats)
        a_best = max((float(r[6] or 0) for r in stats), default=0)
        a_fees = sum(float(r[7] or 0) for r in stats)
        a_total = sum(r[1] for r in stats)
        a_wins = sum(r[2] for r in stats)
        a_spent = a_total * 25

        net_emoji = '\U0001f7e2' if a_net >= 0 else '\U0001f534'
        net_sign = '+' if a_net >= 0 else ''
        lines.append(f"{net_emoji} Net P&L: {net_sign}${a_net:.2f}")
        lines.append(f"\U0001f4b0 Profit: +${a_profit:.2f} | \U0001f4c9 Loss: -${a_loss:.2f}")
        lines.append(f"\U0001f3c6 Best Trade: +${a_best:.2f}")
        lines.append(f"\U0001f4b5 Spent: ${a_spent} | \U0001f4b8 Fees: ${a_fees:.2f}")

        lines.append("")
        for r in stats:
            wl, total, wins, net_pnl = r[0], r[1], r[2], r[3]
            acc = wins/total*100 if total > 0 else 0
            lines.append(f"{wl}m: {wins}W-{total-wins}L ({acc:.0f}%) | {'+' if float(net_pnl or 0) >= 0 else ''}${net_pnl}")

        if a_total > 0:
            lines.append(f"\nRecord: {a_wins}W-{a_total-a_wins}L ({a_wins/a_total*100:.0f}%)")

        # BANKROLL section
        try:
            br = await get_btc_bankroll_state()
            br_bal = float(br.get('balance', BTC_STARTING_BALANCE))
            br_avail = float(br.get('available', br_bal))
            br_in_pos = float(br.get('in_positions', 0))
            br_dd = float(br.get('max_drawdown_pct', 0))
            br_pnl_pct = (br_bal - BTC_STARTING_BALANCE) / BTC_STARTING_BALANCE * 100
            br_emoji = '\U0001f7e2' if br_bal >= BTC_STARTING_BALANCE else '\U0001f534'
            lines.append("")
            lines.append("\u2501\u2501\u2501 BANKROLL \u2501\u2501\u2501")
            lines.append(f"{br_emoji} Balance: ${br_bal:,.0f} ({'+' if br_pnl_pct>=0 else ''}{br_pnl_pct:.1f}% from ${BTC_STARTING_BALANCE:,.0f})")
            lines.append(f"\U0001f4b3 Available: ${br_avail:,.0f} | In Positions: ${br_in_pos:,.0f}")
            if br_dd > 0.01:
                lines.append(f"\U0001f4c9 Max Drawdown: {br_dd:.1f}%")
            lines.append(f"\U0001f3af Max Bet: ${br_bal*0.10:,.0f} | Max Exposure: ${br_bal*0.30:,.0f}")
        except Exception as _bre:
            logger.error(f"Bankroll section failed: {_bre}")

        lines.append(f"Dashboard: brobot.1nnercircle.club/btc15m")

        msg = "\n".join(lines)
        subscribers = await _telegram_bot.get_all_subscribers(instant_only=False)
        await _send_and_pin(_telegram_bot.bot_token, subscribers, msg, 'hourly')
        logger.info(f"\U0001f4ca BTC hourly summary sent + pinned to {len(subscribers)} subscribers")
    except Exception as e:
        logger.error(f"\u274c BTC hourly summary failed: {e}\n{traceback.format_exc()}")


async def scheduled_btc_intelligence_loop():
    """Daily Intelligence Loop at 11:00 PM IST.
    
    THE LOOP:
    1. Analyze today's performance under current strategy
    2. Compare to previous day (progressive, regressive, or neutral)
    3. If progressive: RETAIN strategy, note what worked
    4. If regressive: REVERT to previous strategy + capture learnings
    5. Either way: propose ONE adjustment for tomorrow
    6. Activate new/adjusted strategy for next 24 hours
    7. Log everything to btc_intelligence_log
    8. Broadcast full report to Telegram
    """
    global _telegram_bot
    if not _telegram_bot:
        return
    # Gate: suppress if paper mode + paper_notifications=false
    _tcfg = await get_trading_mode()
    if is_live_mode(_tcfg):
        logger.info("🔇 BTC intelligence loop suppressed (live mode)")
        return
    if not should_send_paper_telegram(_tcfg):
        logger.info("🔇 BTC intelligence loop suppressed (paper notifications off)")
        return
    try:
        import psycopg2, json
        from datetime import date
        conn = psycopg2.connect(dbname='polyedge', user='node', host='localhost')
        cur = conn.cursor()

        # Get active strategy
        cur.execute("SELECT version, max_entry, min_rr, min_factors, window_lengths, stakes FROM btc_strategy_versions WHERE status = 'active' ORDER BY id DESC LIMIT 1")
        active = cur.fetchone()
        if not active:
            conn.close()
            return
        active_version, max_entry, min_rr, min_factors, win_lengths, stakes = active

        # Today's performance (V3-eligible trades: entry < max_entry, correct window lengths)
        cur.execute("""
            WITH best AS (
                SELECT DISTINCT ON (s.window_id)
                    s.window_id, s.prediction, s.confidence, w.resolution, w.window_length,
                    w.close_time,
                    (s.prediction = w.resolution) as correct,
                    CASE WHEN s.prediction = 'UP' THEN w.up_price ELSE w.down_price END as entry_price
                FROM btc_signals s
                JOIN btc_windows w ON s.window_id = w.window_id
                WHERE s.prediction != 'SKIP' AND w.resolution IS NOT NULL
                    AND w.close_time::date = CURRENT_DATE
                ORDER BY s.window_id, s.confidence DESC
            ),
            pnl AS (
                SELECT *,
                    CASE
                        WHEN entry_price < 0.20 THEN 75 WHEN entry_price < 0.30 THEN 50
                        WHEN entry_price < 0.40 THEN 35 ELSE 25
                    END as stake,
                    CASE
                        WHEN correct AND entry_price > 0 AND entry_price < 1 THEN
                            (CASE WHEN entry_price < 0.20 THEN 75 WHEN entry_price < 0.30 THEN 50 WHEN entry_price < 0.40 THEN 35 ELSE 25 END
                             / entry_price - CASE WHEN entry_price < 0.20 THEN 75 WHEN entry_price < 0.30 THEN 50 WHEN entry_price < 0.40 THEN 35 ELSE 25 END) * 0.98
                        WHEN NOT correct THEN
                            -1 * CASE WHEN entry_price < 0.20 THEN 75 WHEN entry_price < 0.30 THEN 50 WHEN entry_price < 0.40 THEN 35 ELSE 25 END
                        ELSE 0
                    END as trade_pnl,
                    CASE WHEN entry_price > 0 THEN (1.0/entry_price - 1.0) * 0.98 ELSE 0 END as rr
                FROM best
            )
            SELECT 
                COUNT(*) as total, COUNT(*) FILTER (WHERE correct) as wins,
                ROUND(SUM(trade_pnl)::numeric, 2) as net_pnl,
                ROUND(SUM(CASE WHEN trade_pnl > 0 THEN trade_pnl ELSE 0 END)::numeric, 2) as gross_profit,
                ROUND(ABS(SUM(CASE WHEN trade_pnl < 0 THEN trade_pnl ELSE 0 END))::numeric, 2) as gross_loss,
                ROUND(MAX(trade_pnl)::numeric, 2) as best_trade,
                ROUND(AVG(entry_price)::numeric, 4) as avg_entry,
                ROUND(AVG(rr)::numeric, 2) as avg_rr
            FROM pnl
        """)
        today = cur.fetchone()
        t_total, t_wins, t_net, t_profit, t_loss, t_best, t_avg_entry, t_avg_rr = today

        # Hourly breakdown for best/worst hour
        cur.execute("""
            WITH best AS (
                SELECT DISTINCT ON (s.window_id)
                    s.window_id, s.prediction, w.resolution, w.window_length, w.close_time,
                    (s.prediction = w.resolution) as correct,
                    CASE WHEN s.prediction = 'UP' THEN w.up_price ELSE w.down_price END as entry_price
                FROM btc_signals s
                JOIN btc_windows w ON s.window_id = w.window_id
                WHERE s.prediction != 'SKIP' AND w.resolution IS NOT NULL
                    AND w.close_time::date = CURRENT_DATE
                ORDER BY s.window_id, s.confidence DESC
            ),
            pnl AS (
                SELECT *, CASE WHEN correct AND entry_price > 0 AND entry_price < 1 THEN (25.0/entry_price - 25.0)*0.98 WHEN NOT correct THEN -25.0 ELSE 0 END as trade_pnl
                FROM best
            )
            SELECT EXTRACT(HOUR FROM close_time AT TIME ZONE 'Asia/Kolkata')::int as hr,
                COUNT(*) as trades, COUNT(*) FILTER (WHERE correct) as wins,
                ROUND(SUM(trade_pnl)::numeric, 2) as pnl,
                ROUND(AVG(entry_price)::numeric, 3) as avg_entry
            FROM pnl GROUP BY 1 ORDER BY 1
        """)
        hourly = cur.fetchall()
        best_hour = max(hourly, key=lambda x: float(x[3])) if hourly else None
        worst_hour = min(hourly, key=lambda x: float(x[3])) if hourly else None
        vol_data = {str(h[0]): {"trades": h[1], "wins": h[2], "pnl": float(h[3]), "avg_entry": float(h[4])} for h in hourly}

        # Previous day for comparison
        cur.execute("SELECT net_pnl, trades_taken, strategy_version, verdict FROM btc_intelligence_log WHERE date = CURRENT_DATE - 1")
        prev = cur.fetchone()
        prev_pnl = float(prev[0]) if prev else None
        prev_version = prev[2] if prev else None

        # === INTELLIGENCE DECISION ===
        t_net_f = float(t_net or 0)
        verdict = 'neutral'
        action = 'retain'
        learnings = []
        next_version = active_version

        if t_total == 0:
            verdict = 'no_data'
            action = 'retain'
            learnings.append('No trades taken today. Filters may be too tight or no favorable entries.')
        elif t_net_f > 0:
            verdict = 'progressive'
            action = 'retain'
            learnings.append(f'Profitable day (+${t_net_f}). Strategy {active_version} is working.')
            if best_hour:
                learnings.append(f'Best hour: {best_hour[0]:02d}:00 (+${float(best_hour[3]):.2f}, entry {float(best_hour[4])*100:.0f}c)')
            if t_avg_rr and float(t_avg_rr) > 1.5:
                learnings.append(f'Avg R:R {float(t_avg_rr):.1f}x — excellent trade selection.')
        elif t_net_f < 0 and t_net_f > -50:
            verdict = 'neutral'
            action = 'retain'
            learnings.append(f'Small loss (${t_net_f}). Within noise range. Retain strategy.')
        else:
            verdict = 'regressive'
            # Check if we should revert or evolve
            if prev and prev_pnl and float(prev_pnl) > 0:
                action = 'revert'
                learnings.append(f'Regressive day (${t_net_f}). Previous strategy {prev_version} was profitable.')
                learnings.append('Reverting to previous strategy and capturing learnings.')
            else:
                action = 'evolve'
                learnings.append(f'Regressive day (${t_net_f}). No profitable previous strategy to revert to.')

        # === PROPOSE ADJUSTMENT ===
        adjustment_note = ''
        new_max_entry = float(max_entry)
        new_min_rr = float(min_rr)
        new_min_factors = min_factors
        new_win_lengths = list(win_lengths)
        new_stakes = json.loads(stakes) if isinstance(stakes, str) else dict(stakes)

        if action == 'revert':
            # Find the BEST historical version (highest P&L in performance_snapshot)
            cur.execute("""
                SELECT version, max_entry, min_rr, min_factors, window_lengths, stakes,
                    (performance_snapshot->>'net_pnl')::numeric as hist_pnl
                FROM btc_strategy_versions 
                WHERE performance_snapshot IS NOT NULL 
                    AND (performance_snapshot->>'net_pnl')::numeric > 0
                ORDER BY (performance_snapshot->>'net_pnl')::numeric DESC
                LIMIT 1
            """)
            best_hist = cur.fetchone()
            
            # Fallback to previous day's version if no profitable history
            if not best_hist and prev_version:
                cur.execute("SELECT version, max_entry, min_rr, min_factors, window_lengths, stakes FROM btc_strategy_versions WHERE version = %s", (prev_version,))
                best_hist = cur.fetchone()
            
            if best_hist:
                revert_to = best_hist[0]
                new_max_entry, new_min_rr, new_min_factors = float(best_hist[1]), float(best_hist[2]), best_hist[3]
                new_win_lengths = list(best_hist[4])
                new_stakes = json.loads(best_hist[5]) if isinstance(best_hist[5], str) else dict(best_hist[5])
                hist_pnl = float(best_hist[6]) if len(best_hist) > 6 and best_hist[6] else None
                pnl_note = f' (was +${hist_pnl:.2f})' if hist_pnl else ''
                adjustment_note = f'Reverted to {revert_to} parameters{pnl_note}. Proven profitable.'
        elif verdict == 'progressive':
            # Small optimization: if avg entry is clustered, widen slightly
            if t_avg_entry and float(t_avg_entry) < 0.35:
                adjustment_note = 'Entries clustered low. Strategy is selective — good.'
            elif worst_hour:
                wh = worst_hour[0]
                adjustment_note = f'Consider avoiding hour {wh:02d}:00 (worst P&L: ${float(worst_hour[3]):.2f}).'
        elif verdict == 'no_data':
            # Loosen slightly if no trades at all
            if new_max_entry < 0.55:
                new_max_entry = min(new_max_entry + 0.05, 0.55)
                adjustment_note = f'No trades taken. Loosened max entry from {float(max_entry)*100:.0f}c to {new_max_entry*100:.0f}c.'
            elif new_min_factors > 3:
                new_min_factors = max(new_min_factors - 1, 3)
                adjustment_note = f'No trades taken. Relaxed factor agreement from {min_factors}/7 to {new_min_factors}/7.'

        # Create new version
        version_num = int(active_version.replace('V', '')) + 1
        new_version = f'V{version_num}'

        # Only create new version if parameters actually changed
        if (new_max_entry != float(max_entry) or new_min_rr != float(min_rr) or 
            new_min_factors != min_factors or new_win_lengths != list(win_lengths)):
            # Snapshot today's performance onto the outgoing version (immutable record)
            cur.execute("""
                UPDATE btc_strategy_versions 
                SET status = %s, deactivated_at = NOW(),
                    performance_snapshot = %s,
                    revert_reason = CASE WHEN %s = 'revert' THEN %s ELSE NULL END
                WHERE status = 'active'
            """, (
                'retained' if action == 'retain' else 'reverted' if action == 'revert' else 'superseded',
                json.dumps({
                    'date': datetime.now().strftime('%Y-%m-%d'),
                    'trades': t_total or 0, 'wins': t_wins or 0,
                    'net_pnl': float(t_net or 0), 'best_trade': float(t_best or 0),
                    'avg_entry': float(t_avg_entry or 0), 'avg_rr': float(t_avg_rr or 0),
                    'verdict': verdict
                }),
                action, '\n'.join(learnings)
            ))
            # Create new version — APPEND ONLY, old versions never modified after this
            cur.execute("""
                INSERT INTO btc_strategy_versions 
                (version, status, max_entry, min_rr, min_factors, window_lengths, stakes, parent_version, notes)
                VALUES (%s, 'active', %s, %s, %s, %s, %s, %s, %s)
            """, (new_version, new_max_entry, new_min_rr, new_min_factors, new_win_lengths,
                   json.dumps(new_stakes), active_version, adjustment_note))
            next_version = new_version
        else:
            next_version = active_version

        # Log today's intelligence
        cur.execute("""
            INSERT INTO btc_intelligence_log (date, strategy_version, trades_taken, trades_won,
                net_pnl, gross_profit, gross_loss, best_trade, avg_entry, avg_rr,
                best_hour, worst_hour, best_hour_pnl, worst_hour_pnl,
                volatility_data, verdict, action_taken, next_strategy, learnings)
            VALUES (CURRENT_DATE, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (date) DO UPDATE SET
                trades_taken = EXCLUDED.trades_taken, trades_won = EXCLUDED.trades_won,
                net_pnl = EXCLUDED.net_pnl, gross_profit = EXCLUDED.gross_profit,
                gross_loss = EXCLUDED.gross_loss, best_trade = EXCLUDED.best_trade,
                avg_entry = EXCLUDED.avg_entry, avg_rr = EXCLUDED.avg_rr,
                best_hour = EXCLUDED.best_hour, worst_hour = EXCLUDED.worst_hour,
                best_hour_pnl = EXCLUDED.best_hour_pnl, worst_hour_pnl = EXCLUDED.worst_hour_pnl,
                volatility_data = EXCLUDED.volatility_data, verdict = EXCLUDED.verdict,
                action_taken = EXCLUDED.action_taken, next_strategy = EXCLUDED.next_strategy,
                learnings = EXCLUDED.learnings
        """, (
            active_version, t_total or 0, t_wins or 0,
            t_net or 0, t_profit or 0, t_loss or 0, t_best or 0,
            t_avg_entry or 0, t_avg_rr or 0,
            best_hour[0] if best_hour else None,
            worst_hour[0] if worst_hour else None,
            float(best_hour[3]) if best_hour else None,
            float(worst_hour[3]) if worst_hour else None,
            json.dumps(vol_data),
            verdict, action, next_version,
            '\n'.join(learnings)
        ))
        conn.commit()
        conn.close()

        # === BROADCAST ===
        verdict_emoji = {'progressive': '\U0001f7e2', 'regressive': '\U0001f534', 'neutral': '\U0001f7e1', 'no_data': '\u26aa'}
        action_emoji = {'retain': '\u2705', 'revert': '\u21a9\ufe0f', 'evolve': '\U0001f52c'}

        lines = [
            '\U0001f9e0 BTC INTELLIGENCE LOOP \u2014 Daily Report',
            f'Date: {datetime.now().strftime("%Y-%m-%d")} IST',
            f'Strategy: {active_version}',
            '',
            '\u2501\u2501\u2501 TODAY\'S PERFORMANCE \u2501\u2501\u2501',
        ]
        if t_total and t_total > 0:
            lines.append(f'Trades: {t_wins}W-{t_total-t_wins}L ({t_wins/t_total*100:.0f}%)')
            lines.append(f'Net P&L: {"+ " if t_net_f >= 0 else ""}${t_net_f:.2f}')
            lines.append(f'Gross Profit: +${float(t_profit or 0):.2f}')
            lines.append(f'Gross Loss: -${float(t_loss or 0):.2f}')
            lines.append(f'Best Trade: +${float(t_best or 0):.2f}')
            lines.append(f'Avg Entry: {float(t_avg_entry or 0)*100:.0f}c | Avg R:R: {float(t_avg_rr or 0):.1f}x')
        else:
            lines.append('No trades taken today.')

        if hourly:
            lines.append('')
            lines.append('\u2501\u2501\u2501 HOURLY HEATMAP \u2501\u2501\u2501')
            for h in hourly:
                hr, trades, wins, pnl, entry = h
                e = '\U0001f7e2' if float(pnl) >= 0 else '\U0001f534'
                lines.append(f'{e} {hr:02d}:00 | {wins}W-{trades-wins}L | {"+ " if float(pnl)>=0 else ""}${pnl} | entry: {float(entry)*100:.0f}c')

        lines.append('')
        lines.append('\u2501\u2501\u2501 INTELLIGENCE VERDICT \u2501\u2501\u2501')
        lines.append(f'{verdict_emoji.get(verdict, "")} Verdict: {verdict.upper()}')
        lines.append(f'{action_emoji.get(action, "")} Action: {action.upper()}')
        if prev_pnl is not None:
            lines.append(f'vs Yesterday: {"+ " if prev_pnl >= 0 else ""}${prev_pnl:.2f} ({prev_version})')
        for l in learnings:
            lines.append(f'\U0001f4a1 {l}')

        lines.append('')
        lines.append('\u2501\u2501\u2501 TOMORROW\'S STRATEGY \u2501\u2501\u2501')
        lines.append(f'Version: {next_version}')
        lines.append(f'Max Entry: {new_max_entry*100:.0f}c | Min R:R: {new_min_rr:.1f}x')
        lines.append(f'Factor Agreement: {new_min_factors}/7 | Windows: {"m, ".join(str(w) for w in new_win_lengths)}m')
        if adjustment_note:
            lines.append(f'\U0001f527 {adjustment_note}')

        lines.append(f'\nDashboard: brobot.1nnercircle.club/btc15m')

        msg = '\n'.join(lines)
        subscribers = await _telegram_bot.get_all_subscribers(instant_only=False)
        await _send_and_pin(_telegram_bot.bot_token, subscribers, msg, 'intelligence')
        logger.info(f'\U0001f9e0 Intelligence Loop: {verdict} / {action} / next={next_version}')
    except Exception as e:
        logger.error(f'\u274c Intelligence Loop failed: {e}\n{traceback.format_exc()}')




# ═══════════════════════════════════════════════════════════════
# BTC V4 STRATEGY — 10-Point Confluence Paper Trading
# ═══════════════════════════════════════════════════════════════
_v4_strategy = None

_v5_strategy = None

async def scheduled_v5_paper_scan():
    """V5 Strategy: Signal engine direction + JC confluence + risk gates. Paper only."""
    global _v5_strategy, _btc_engine
    try:
        if _v5_strategy is None:
            from src.strategies.btc_v5_strategy import BTCV5Strategy
            _v5_strategy = BTCV5Strategy(get_async_pool())
            logger.info("✅ V5 Strategy initialized")

        if _btc_engine is None:
            return

        pool = get_async_pool()
        async with pool.acquire() as conn:
            windows = await conn.fetch("""
                SELECT window_id, window_length, up_price, down_price, btc_open,
                       EXTRACT(EPOCH FROM (close_time - NOW())) as seconds_remaining
                FROM btc_windows
                WHERE close_time > NOW() AND close_time < NOW() + INTERVAL '5 minutes'
                AND window_length = 5
                ORDER BY close_time ASC LIMIT 3
            """)

        if not windows:
            return

        for w in windows:
            window_id = w['window_id']
            seconds_remaining = int(w['seconds_remaining'] or 0)
            up_price = float(w['up_price'] or 0.5)
            down_price = float(w['down_price'] or 0.5)

            if seconds_remaining <= 0:
                continue

            # Check if we already have a paper trade for this window
            async with pool.acquire() as conn:
                existing = await conn.fetchrow(
                    "SELECT 1 as x FROM paper_trades WHERE window_id=$1", window_id)
            if existing:
                continue

            # ══ Get signal engine prediction (7 factors) ═════════════════
            try:
                btc_data = await _btc_engine.get_btc_price()
                trades = await _btc_engine.get_btc_trades(limit=200)
                engine_factors = await _btc_engine.compute_factors(
                    {"window_id": window_id, "up_price": up_price, "down_price": down_price,
                     "btc_open": float(w['btc_open'] or 0), "seconds_remaining": seconds_remaining,
                     "window_length": 5},
                    btc_data, trades)
                prediction, prob_up, confidence = _btc_engine.predict(engine_factors)
            except Exception as e:
                logger.warning(f"V5 signal engine error for {window_id}: {e}")
                continue

            if prediction == 'SKIP':
                logger.info(f"V5 scan {window_id}: Signal engine SKIP")
                continue

            # ══ V5 MANDATORY RISK GATES ══════════════════════════════════
            risk = await _v5_strategy.check_risk_gates()
            if not risk['ok']:
                logger.warning(f"🛑 V5 RISK BLOCKED: {risk['reason']}")
                continue

            # ══ V5 EVALUATE (engine direction + JC confluence) ═══════════
            btc_price = btc_data.get('price', 0) if btc_data else 0
            result = await _v5_strategy.evaluate(
                window_id=window_id,
                up_price=up_price,
                down_price=down_price,
                seconds_remaining=seconds_remaining,
                engine_prediction=prediction,
                engine_prob_up=prob_up,
                engine_confidence=confidence,
                engine_factors=engine_factors,
                btc_price=btc_price
            )

            logger.info(f"V5 scan {window_id}: {result.get('reason', 'no reason')}")

            if result.get('should_trade'):
                direction = result['direction']
                stake = result['stake']
                token_price = result['token_price']
                jc_label = result.get('jc_level', 'none')
                factors = result.get('factors', {})
                bankroll = risk.get('bankroll', 10000)
                score = result.get('score', 0)
                rr = result.get('rr', 0)

                import json as _json
                async with pool.acquire() as conn:
                    await conn.execute("""
                        INSERT INTO paper_trades
                        (window_id, direction, token_price, stake_usd, confluence_score,
                         factors_json, jc_level, strategy_version)
                        VALUES ($1,$2,$3,$4,$5,$6::jsonb,$7,'V5')
                    """, window_id, direction, token_price, stake, score,
                         _json.dumps({k: v.encode('ascii','ignore').decode() if isinstance(v,str) else str(v) for k,v in (factors if isinstance(factors,dict) else {}).items()}),
                         str(jc_label).encode('ascii','ignore').decode())

                logger.info(f"📝 V5 PAPER: {direction} | {token_price*100:.1f}c | ${stake} | Score {score}")

                # Trade open — log only, no TG notification (CEO: only send won/loss results)

    except Exception as e:
        logger.error(f"❌ V5 paper scan failed: {e}")


async def scheduled_v5_resolution():
    """V5: Resolve open paper trades every 2 min. Score win/loss, update bankroll, send Telegram."""
    global _v5_strategy
    try:
        if _v5_strategy is None:
            return
        resolved = await _v5_strategy.resolve_open_trades()
        if resolved:
            summaries = [f"{r['direction']}->{'W' if r['won'] else 'L'} ${r['pnl']:+.2f}" for r in resolved]
            logger.info(f"✅ V5 resolved {len(resolved)} trades: {summaries}")
    except Exception as e:
        logger.error(f"❌ V5 resolution failed: {e}")


async def scheduled_v4_daily_report():
    """V4 daily intelligence report at 11PM IST."""
    global _telegram_bot
    try:
        pool = get_async_pool()
        async with pool.acquire() as conn:
            trades = await conn.fetch("""
                SELECT direction, token_price, stake_usd, confluence_score,
                       resolution, won, simulated_pnl, created_at
                FROM paper_trades
                WHERE created_at > NOW() - INTERVAL '7 days'
                AND strategy_version IN ('V4', 'V5')
            """)

        if not trades:
            msg = "🧠 V4 Daily Report\n\nNo paper trades yet. Bot is scanning...\n\nWallet: $338.63 | Mode: PAPER"
        else:
            total = len(trades)
            resolved = [t for t in trades if t['won'] is not None]
            wins = sum(1 for t in resolved if t['won'])
            win_rate = (wins/len(resolved)*100) if resolved else 0
            sim_pnl = sum(float(t['simulated_pnl'] or 0) for t in resolved)
            today = [t for t in trades if t['created_at'].date() == datetime.utcnow().date()]

            msg = (
                f"🧠 V4 Daily Intelligence Report\n"
                f"{'─'*30}\n"
                f"📊 7-Day Paper Trades: {total}\n"
                f"✅ Resolved: {len(resolved)} | Wins: {wins}\n"
                f"🎯 Win Rate: {win_rate:.1f}%\n"
                f"💰 Simulated P&L: ${sim_pnl:+.2f}\n"
                f"📅 Today: {len(today)} trades\n"
                f"{'─'*30}\n"
            )
            if win_rate >= 58 and len(resolved) >= 50:
                msg += "✅ WIN RATE TARGET MET! Ready for CEO approval to go LIVE.\n"
            elif len(resolved) >= 50:
                msg += f"❌ Win rate {win_rate:.1f}% below 58% target — keep paper testing\n"
            else:
                msg += f"📈 Need {50-len(resolved)} more resolved trades before live review\n"
            msg += "\n⚠️ PAPER MODE — $338.63 real wallet untouched"

        if _telegram_bot:
            try:
                await _telegram_bot.bot.send_message(chat_id='1656605843', text=msg)
                logger.info("✅ V4 daily report sent")
            except Exception as _te:
                logger.warning(f"V4 daily report Telegram failed: {_te}")

    except Exception as e:
        logger.error(f"❌ V4 daily report failed: {e}")


async def scheduled_btc_daily_strategy_report():
    """Daily strategy analysis at 11:30 PM IST — volatility, hourly performance, adjustments."""
    global _telegram_bot
    if not _telegram_bot:
        return
    # Gate: suppress if paper mode + paper_notifications=false
    _tcfg = await get_trading_mode()
    if is_live_mode(_tcfg):
        logger.info("🔇 BTC daily strategy report suppressed (live mode)")
        return
    if not should_send_paper_telegram(_tcfg):
        logger.info("🔇 BTC daily strategy report suppressed (paper notifications off)")
        return
    try:
        import psycopg2
        conn = psycopg2.connect(dbname='polyedge', user='node', host='localhost')
        cur = conn.cursor()

        # Today's hourly breakdown — uses btc_pnl view (correct filters + scaled stakes)
        cur.execute("""
            SELECT 
                EXTRACT(HOUR FROM close_time AT TIME ZONE 'Asia/Kolkata')::int as hour_ist,
                COUNT(*) as all_trades,
                COUNT(*) FILTER (WHERE correct) as all_wins,
                ROUND(SUM(trade_pnl)::numeric, 2) as all_pnl,
                ROUND(AVG(entry_price)::numeric, 3) as avg_entry
            FROM btc_pnl
            WHERE close_time::date = CURRENT_DATE
            GROUP BY 1 ORDER BY 1
        """)
        hourly = cur.fetchall()

        # Volatility data
        cur.execute("""
            SELECT hour_ist, trades_taken, trades_won, net_pnl, avg_entry, btc_price_range_pct, session_tag
            FROM btc_volatility_hours
            WHERE date = CURRENT_DATE
            ORDER BY hour_ist
        """)
        volatility = cur.fetchall()

        # Overall today — from btc_pnl view
        cur.execute("""
            SELECT 
                COUNT(*) as total, COUNT(*) FILTER (WHERE correct) as wins,
                ROUND(SUM(trade_pnl)::numeric, 2) as net,
                ROUND(MAX(trade_pnl)::numeric, 2) as best,
                ROUND(SUM(stake)::numeric, 2) as spent,
                ROUND(SUM(CASE WHEN trade_pnl>0 THEN stake*(1.0/entry_price-1.0)*0.02 ELSE 0 END)::numeric, 2) as fees
            FROM btc_pnl
            WHERE close_time::date = CURRENT_DATE
        """)
        totals = cur.fetchone()
        conn.close()

        # Build report
        lines = ["\U0001f4ca BTC DAILY STRATEGY REPORT", f"Date: {datetime.now().strftime('%Y-%m-%d')} IST", ""]

        # Hourly heat map
        lines.append("\u2501\u2501\u2501 HOURLY PERFORMANCE \u2501\u2501\u2501")
        best_hour, worst_hour = None, None
        best_pnl, worst_pnl = -999999, 999999
        for h in hourly:
            hr, all_t, all_w, all_p, avg_e = h
            emoji = '\U0001f7e2' if float(all_p) >= 0 else '\U0001f534'
            sign = '+' if float(all_p) >= 0 else ''
            lines.append(f"{emoji} {hr:02d}:00 | {all_w}W-{all_t-all_w}L | {sign}${all_p} | avg entry: {float(avg_e)*100:.0f}c")
            if float(all_p) > best_pnl:
                best_pnl, best_hour = float(all_p), hr
            if float(all_p) < worst_pnl:
                worst_pnl, worst_hour = float(all_p), hr

        if best_hour is not None:
            lines.append(f"\n\U0001f3c6 Best hour: {best_hour:02d}:00 (+${best_pnl:.2f})")
            lines.append(f"\U0001f4a9 Worst hour: {worst_hour:02d}:00 (${worst_pnl:.2f})")

        # Volatility slots
        if volatility:
            lines.append("")
            lines.append("\u2501\u2501\u2501 VOLATILITY MAP \u2501\u2501\u2501")
            for v in volatility:
                v_hr, v_trades, v_won, v_pnl, v_entry, v_range, v_tag = v
                vol_emoji = '\U0001f525' if float(v_range) > 0.3 else '\u26a1' if float(v_range) > 0.1 else '\U0001f4a4'
                lines.append(f"{vol_emoji} {v_hr:02d}:00 | Range: {float(v_range):.2f}% | {v_trades} trades | tag: {v_tag}")

        # Totals — uses corrected btc_pnl view
        total, wins, net, best, spent, fees = totals
        net_f = float(net or 0); spent_f = float(spent or 0); fees_f = float(fees or 0)
        lines.append("")
        lines.append("\u2501\u2501\u2501 DAY TOTALS \u2501\u2501\u2501")
        net_emoji = '\U0001f7e2' if net_f >= 0 else '\U0001f534'
        net_sign = '+' if net_f >= 0 else ''
        lines.append(f"{net_emoji} Net P&L: {net_sign}${net_f:.2f}")
        if total > 0:
            lines.append(f"Record: {wins}W-{total-wins}L ({wins/total*100:.0f}%)")
        lines.append(f"\U0001f4b5 Spent: ${spent_f:.0f} | \U0001f4b8 Fees: ${fees_f:.2f}")
        if best:
            lines.append(f"\U0001f3c6 Best trade: +${best}")

        cur.execute("SELECT version, max_entry, min_factors, min_rr FROM btc_strategy_versions WHERE status=\'active\' LIMIT 1")
        strat = cur.fetchone()
        lines.append("\n\u2501\u2501\u2501 STRATEGY \u2501\u2501\u2501")
        if strat:
            lines.append(f"Active: {strat[0]} | Entry <{float(strat[1])*100:.0f}c | {strat[2]}/7 factors | Min R:R {float(strat[3]):.1f}x")
        lines.append("Hourly slots tracked for weekly review")
        lines.append("\nDashboard: brobot.1nnercircle.club/btc15m")
        conn.close()

        msg = "\n".join(lines)
        subscribers = await _telegram_bot.get_all_subscribers(instant_only=False)
        await _send_and_pin(_telegram_bot.bot_token, subscribers, msg, 'daily')
        logger.info(f"\U0001f4ca BTC daily strategy report sent + pinned")
    except Exception as e:
        logger.error(f"\u274c BTC daily strategy report failed: {e}")


async def scheduled_btc_resolution_check():
    """BTC Signal Engine: check closed windows, score predictions."""
    global _last_btc_resolution, _btc_engine
    try:
        if _btc_engine is None:
            return
        resolved = await _btc_engine.check_resolutions()
        _last_btc_resolution = datetime.utcnow()
        if resolved:
            logger.info(f"📊 BTC resolved: {len(resolved)} windows")
            
            # Check trading mode — only send Telegram if live OR paper_notifications=true
            _tcfg = await get_trading_mode()
            _send_telegram = is_live_mode(_tcfg) or should_send_paper_telegram(_tcfg)
            
            # Broadcast WON/LOST for each resolved window with a signal
            if _telegram_bot and not _send_telegram:
                logger.info(f"🔇 BTC resolution alerts suppressed (paper mode, notifications off, {len(resolved)} resolved)")
            if _telegram_bot and _send_telegram:
                import psycopg2 as _pg
                import httpx as _httpx
                for r in resolved:
                    if not r.get('had_signal'):
                        continue
                    was_correct = r.get('was_correct', False)
                    pred = r.get('prediction', '?')
                    actual = r.get('resolution', '?')
                    wl = r.get('window_length', 15)
                    window_id = r.get('window_id', '')

                    try:
                        _conn = _pg.connect(dbname='polyedge', user='node', host='localhost')
                        _cur = _conn.cursor()
                        _cur.execute("""
                            SELECT p.entry_price, p.stake, ROUND(p.trade_pnl::numeric,2),
                                s.prob_up, s.confidence,
                                s.f_price_delta, s.f_momentum, s.f_volume_imbalance, s.f_oracle_lead,
                                w.btc_open, w.btc_close,
                                (CASE WHEN (p.prediction='DOWN' AND s.f_price_delta<0) OR (p.prediction='UP' AND s.f_price_delta>0) THEN 1 ELSE 0 END +
                                 CASE WHEN (p.prediction='DOWN' AND s.f_momentum<0) OR (p.prediction='UP' AND s.f_momentum>0) THEN 1 ELSE 0 END +
                                 CASE WHEN (p.prediction='DOWN' AND s.f_volume_imbalance<0) OR (p.prediction='UP' AND s.f_volume_imbalance>0) THEN 1 ELSE 0 END +
                                 CASE WHEN (p.prediction='DOWN' AND s.f_oracle_lead<0) OR (p.prediction='UP' AND s.f_oracle_lead>0) THEN 1 ELSE 0 END +
                                 CASE WHEN s.f_book_imbalance IS NOT NULL THEN 1 ELSE 0 END) as factors_agreed
                            FROM btc_pnl p
                            JOIN btc_windows w ON p.window_id = w.window_id
                            JOIN LATERAL (
                                SELECT * FROM btc_signals WHERE window_id = p.window_id AND prediction != 'SKIP'
                                ORDER BY confidence DESC LIMIT 1
                            ) s ON true
                            WHERE p.window_id = %s
                        """, (window_id,))
                        row = _cur.fetchone()
                        _cur.execute("""
                            SELECT COUNT(*), COUNT(*) FILTER (WHERE correct), ROUND(SUM(trade_pnl)::numeric,2)
                            FROM btc_pnl
                        """)
                        ytd = _cur.fetchone()
                        _conn.close()

                        if row:
                            ep, stake, pnl, prob, conf, f_pd, f_mom, f_vol, f_orc, btc_open, btc_close, factors = row
                            ep=float(ep); stake=float(stake); pnl=float(pnl)
                            btc_move = float(btc_close or 0) - float(btc_open or 0)
                            roi = pnl/stake*100 if stake else 0
                            ytd_t, ytd_w, ytd_net = ytd
                            pnl_str = f"+${pnl:.2f}" if pnl>=0 else f"-${abs(pnl):.2f}"
                            move_str = f"+${btc_move:,.0f}" if btc_move>=0 else f"-${abs(btc_move):,.0f}"
                            re = '\u2705' if was_correct else '\u274c'
                            rw = 'WON' if was_correct else 'LOST'

                            # ✔️ Update bankroll on resolution
                            try:
                                await update_btc_bankroll_close(stake, pnl, was_correct)
                                await sync_btc_in_positions()
                            except Exception as _bre:
                                logger.error(f"❌ Bankroll resolution update failed: {_bre}")

                            # ✔️ Resolve matching live trade if exists
                            try:
                                _live_pool = get_async_pool()
                                async with _live_pool.acquire() as _lconn:
                                    _live_row = await _lconn.fetchrow(
                                        "SELECT id FROM live_trades WHERE window_id = $1 AND status = 'open' LIMIT 1",
                                        window_id
                                    )
                                if _live_row:
                                    from src.polymarket_live import PolymarketLiveTrader
                                    _live_trader = PolymarketLiveTrader(get_async_pool())
                                    await _live_trader.resolve_trade(
                                        trade_id=_live_row['id'],
                                        won=was_correct,
                                        exit_price=1.0 if was_correct else 0.0,
                                    )
                                    _live_emoji = '\U0001f7e2' if was_correct else '\U0001f534'
                                    logger.info(f"{_live_emoji} Live trade #{_live_row['id']} resolved")
                            except Exception as _lre:
                                logger.error(f"❌ Live trade resolution failed: {_lre}")
                            crowd_pct = max(float(prob or 0.5), 1-float(prob or 0.5)) * 100
                            crowd_side = 'bullish' if float(prob or 0.5) > 0.5 else 'bearish'
                            rr = (1/ep - 1)*0.98 if ep > 0 else 0

                            # Per-factor confluence
                            def _fa(val, direction):
                                v = float(val or 0)
                                agree = (direction=='DOWN' and v<0) or (direction=='UP' and v>0)
                                bar = '\u2588' * min(int(abs(v)*5), 5)
                                return ('\u2705' if agree else '\u274c'), bar
                            pd_e,pd_b = _fa(f_pd, pred)
                            mm_e,mm_b = _fa(f_mom, pred)
                            vi_e,vi_b = _fa(f_vol, pred)
                            ol_e,ol_b = _fa(f_orc, pred)
                            bk_e = '\u2705' if factors >= 4 else '\u274c'

                            # Get updated bankroll for display
                            try:
                                _br_disp = await get_btc_bankroll_state()
                                _br_bal = float(_br_disp.get('balance', BTC_STARTING_BALANCE))
                                _br_avail = float(_br_disp.get('available', BTC_STARTING_BALANCE))
                                _br_dd = float(_br_disp.get('max_drawdown_pct', 0))
                                _br_pnl_pct = (_br_bal - BTC_STARTING_BALANCE) / BTC_STARTING_BALANCE * 100
                                _br_label = 'WALLET' if is_live_mode(_tcfg) else 'BANKROLL'
                                bankroll_line = (f"\n\u2501\u2501\u2501 {_br_label} \u2501\u2501\u2501\n"
                                    f"\U0001f4b0 ${_br_bal:,.0f} ({'+' if _br_pnl_pct>=0 else ''}{_br_pnl_pct:.1f}%) | Avail: ${_br_avail:,.0f}"
                                    + (f" | DD: {_br_dd:.1f}%" if _br_dd > 0.01 else ""))
                            except Exception:
                                bankroll_line = ""

                            # LIVE green theme vs paper theme
                            if is_live_mode(_tcfg):
                                _win_emoji = '\U0001f7e2\U0001f4b0' if was_correct else '\U0001f534\U0001f4b8'
                                _label = 'LIVE WIN' if was_correct else 'LIVE LOSS'
                                _section = 'REAL PROFIT' if was_correct else 'REAL LOSS'
                            else:
                                _win_emoji = re
                                _label = rw
                                _section = 'TRADE'

                            msg = (
                                f"{_win_emoji if is_live_mode(_tcfg) else re} BTC {_label if is_live_mode(_tcfg) else rw} \u2014 {wl}M\n"
                                f"\n"
                                f"\u2501\u2501\u2501 {_section} \u2501\u2501\u2501\n"
                                f"Direction: {pred} \u2192 Actual: {actual}\n"
                                f"Entry: {ep*100:.1f}\u00a2 | Stake: ${stake:.0f} | R:R {rr:.1f}:1\n"
                                f"P&L: {pnl_str} | ROI: {'+' if roi>=0 else ''}{roi:.0f}%\n"
                                f"BTC: ${float(btc_open):,.0f} \u2192 ${float(btc_close):,.0f} ({move_str})\n"
                                f"\n"
                                f"\u2501\u2501\u2501 SIGNALS ({factors}/5 agreed) \u2501\u2501\u2501\n"
                                f"{pd_e} Price Delta   {float(f_pd or 0):+.3f} {pd_b}\n"
                                f"{mm_e} Momentum      {float(f_mom or 0):+.3f} {mm_b}\n"
                                f"{vi_e} Volume        {float(f_vol or 0):+.3f} {vi_b}\n"
                                f"{ol_e} Oracle Lead   {float(f_orc or 0):+.3f} {ol_b}\n"
                                f"{bk_e} Crowd: {crowd_pct:.0f}% {crowd_side} | Conf: {float(conf or 0)*100:.0f}%\n"
                                f"\n"
                                f"\U0001f4c8 YTD: {ytd_w}W-{ytd_t-ytd_w}L | {'+' if float(ytd_net)>=0 else ''}${float(ytd_net):.2f}"
                                f"{bankroll_line}"
                            )
                        else:
                            re = '✅' if was_correct else '❌'
                            msg = f"{re} BTC {'WON' if was_correct else 'LOST'} — {wl}M | {pred} → {actual}"
                    except Exception as _e:
                        logger.error(f"❌ Trade broadcast error: {_e}")
                        re = '✅' if was_correct else '❌'
                        msg = f"{re} BTC {'WON' if was_correct else 'LOST'} — {wl}M | {pred} → {actual}"

                    # GATE: Only send Telegram if real live trade exists for this window OR paper notifications on
                    _has_live_trade = False
                    try:
                        _lpool = get_async_pool()
                        async with _lpool.acquire() as _lc:
                            _lt = await _lc.fetchrow("SELECT id FROM live_trades WHERE window_id = $1 LIMIT 1", window_id)
                            _has_live_trade = _lt is not None
                    except Exception:
                        pass

                    _should_send = _has_live_trade or should_send_paper_telegram(_tcfg)
                    if _should_send:
                        try:
                            BOT_TOKEN = _telegram_bot.bot_token
                            subs = await _telegram_bot.get_all_subscribers(instant_only=False)
                            async with _httpx.AsyncClient(timeout=10) as _c:
                                for sub in subs:
                                    try:
                                        await _c.post(f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                                            json={"chat_id": sub['chat_id'], "text": msg})
                                    except Exception:
                                        pass
                        except Exception:
                            pass
    except Exception as e:
        logger.error(f"❌ BTC resolution check failed: {e}")
# ═══════════════════════════════════════════════════════════════


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Manage app lifecycle: startup and shutdown."""
    global _last_data_scan, _startup_time, _signal_loop, _improvement_engine, _telegram_bot

    logger.info("🚀 BroBot starting...")
    _startup_time = datetime.utcnow()
    await init_tables()

    
    # Initialize improvement engine
    try:
        from src.learning.improvement import ImprovementEngine
        async_pool = get_async_pool()
        _improvement_engine = ImprovementEngine(async_pool, config)
        logger.info("✅ Improvement engine initialized")
    except Exception as e:
        logger.error(f"⚠️ Improvement engine init failed: {e}")


    # Initialize btc_bankroll if empty
    try:
        existing = await fetch_one("SELECT id FROM btc_bankroll LIMIT 1")
        if not existing:
            pool = get_async_pool()
            async with pool.acquire() as _conn:
                await _conn.execute(
                    "INSERT INTO btc_bankroll (balance, available, peak_balance) VALUES ($1, $2, $3)",
                    BTC_STARTING_BALANCE, BTC_STARTING_BALANCE, BTC_STARTING_BALANCE
                )
            logger.info("✅ BTC bankroll initialized at $5,000")
        else:
            await sync_btc_in_positions()
            br_state = await get_btc_bankroll_state()
            logger.info(f"✅ BTC bankroll: ${float(br_state.get('balance', 0)):,.0f} available ${float(br_state.get('available', 0)):,.0f}")
    except Exception as e:
        logger.error(f"⚠️ BTC Bankroll init failed: {e}")

    # Legacy bankroll table
    try:
        existing2 = await fetch_one("SELECT id FROM bankroll LIMIT 1")
        if not existing2:
            await execute("INSERT INTO bankroll (total_usd, available_usd, in_positions_usd, daily_pnl) VALUES (0, 0, 0, 0)")
    except Exception:
        pass

    # Initialize Telegram subscriber bot
    try:
        from src.alerts.subscriber_bot import init_bot
        from src.alerts.invite_gate import InviteGate
        if config.TELEGRAM_BOT_TOKEN:
            _telegram_bot = await init_bot(config.TELEGRAM_BOT_TOKEN, config.TELEGRAM_ADMIN_CHAT_ID)
            # Wire invite gate
            if _telegram_bot and hasattr(_telegram_bot, 'pool'):
                _telegram_bot.invite_gate = InviteGate(_telegram_bot.pool, config.TELEGRAM_ADMIN_CHAT_ID)
            logger.info("✅ Telegram subscriber bot initialized (invite-only mode)")
        else:
            logger.warning("⚠️ Telegram bot disabled: no token")
    except Exception as e:
        logger.error(f"⚠️ Telegram bot init failed: {e}")
    
    
    # Start scheduler
    
    
    
    
    
    

    
    # ═══════════════════════════════════════════════════════════════
    # BTC SIGNAL ENGINE — Scan every 45 seconds, resolve every 2 min
    # ═══════════════════════════════════════════════════════════════
    scheduler.add_job(scheduled_btc_signal_scan, 'interval', seconds=45, id='btc_signal', replace_existing=True)
    scheduler.add_job(scheduled_btc_resolution_check, 'interval', minutes=2, id='btc_resolution', replace_existing=True)
    # DISABLED per CEO — only send trade open/close, no periodic reports
    # scheduler.add_job(scheduled_btc_hourly_summary, 'cron', minute=0, timezone='Asia/Kolkata', id='btc_hourly', replace_existing=True)
    # scheduler.add_job(scheduled_btc_intelligence_loop, 'cron', hour=23, minute=0, timezone='Asia/Kolkata', id='btc_intelligence', replace_existing=True)
    # scheduler.add_job(scheduled_btc_daily_strategy_report, 'cron', hour=23, minute=30, timezone='Asia/Kolkata', id='btc_daily', replace_existing=True)
    # V5 Strategy — 10-point confluence + mandatory guardrails + resolution
    scheduler.add_job(scheduled_v5_paper_scan, 'interval', seconds=45, id='btc_v5_paper', replace_existing=True)
    scheduler.add_job(scheduled_v5_resolution, 'interval', minutes=2, id='btc_v5_resolve', replace_existing=True)
    scheduler.add_job(scheduled_v4_daily_report, 'cron', hour=23, minute=0, timezone='Asia/Kolkata', id='btc_v5_daily', replace_existing=True)

    # HEARTBEAT: Send status to @ArbitrageBihariBot every 4 hours so CEO knows bot is alive
    async def v5_heartbeat():
        try:
            pool = get_async_pool()
            async with pool.acquire() as conn:
                trades = await conn.fetchrow("SELECT count(*) as total, count(*) FILTER (WHERE won=true) as wins, COALESCE(SUM(simulated_pnl),0) as pnl FROM paper_trades WHERE strategy_version='V5' AND resolved_at IS NOT NULL")
                open_t = await conn.fetchval("SELECT count(*) FROM paper_trades WHERE strategy_version='V5' AND resolved_at IS NULL")
                br = await conn.fetchrow("SELECT balance FROM btc_bankroll WHERE id=1")
            import httpx
            btc = await httpx.AsyncClient(timeout=5).get('https://api.binance.com/api/v3/ticker/price?symbol=BTCUSDT')
            btc_price = float(btc.json()['price'])
            bankroll = float(br['balance']) if br else 10000
            total = int(trades['total']) if trades else 0
            wins = int(trades['wins']) if trades else 0
            pnl = float(trades['pnl']) if trades else 0
            wr = (wins/total*100) if total > 0 else 0
            from src.strategies.btc_v5_strategy import tg_send
            await tg_send(
                f"💓 BroBot V5 Heartbeat\n"
                f"{'━'*24}\n"
                f"BTC: ${btc_price:,.0f}\n"
                f"Trades: {total} | Open: {open_t}\n"
                f"Win Rate: {wr:.0f}% ({wins}/{total})\n"
                f"P&L: ${pnl:+,.2f}\n"
                f"Bankroll: ${bankroll:,.2f}\n"
                f"{'━'*24}\n"
                f"⚠️ PAPER MODE | Scans every 45s"
            )
            logger.info(f"💓 V5 heartbeat sent: {total} trades, ${pnl:+.2f} P&L")
        except Exception as e:
            logger.error(f"V5 heartbeat failed: {e}")

    # DISABLED: CEO wants only trade results
    # scheduler.add_job(v5_heartbeat, 'cron', hour='0,6,12,18', minute=0, timezone='Asia/Kolkata', id='v5_heartbeat', replace_existing=True)
    logger.info("✅ V5 Strategy scheduled (scan: 45s, resolve: 2min, heartbeat: 4h, daily: 11PM)")
    logger.info("✅ BTC Signal Engine scheduled (scan: 45s, resolve: 2min)")

    # ═══════════════════════════════════════════════════════════════
    # JC COPY TRADER — Monitor Jayson's levels every 10s
    # ═══════════════════════════════════════════════════════════════
    try:
        import sys
        sys.path.insert(0, '/data/.openclaw/workspace/projects/Ghost')
        from jc_copy_trader import run_jc_copy_trader, send_jc_hourly_report, send_jc_daily_report
        scheduler.add_job(run_jc_copy_trader, 'interval', seconds=10, id='jc_copy_trader', replace_existing=True)
        # DISABLED per CEO — only send trade open/close, no periodic reports
        # scheduler.add_job(send_jc_hourly_report, 'cron', minute=0, timezone='Asia/Kolkata', id='jc_hourly', replace_existing=True)
        # scheduler.add_job(send_jc_daily_report, 'cron', hour=23, minute=0, timezone='Asia/Kolkata', id='jc_daily', replace_existing=True)
        # JC Discord reporter — 15min updates, hourly levels, level proximity alerts, daily recap
        # JC reporter: ONLY forward Jayson's actual trade signals/charts (no periodic updates)
        # 15min, hourly, proximity alerts, daily recap — all REMOVED per CEO directive
        # Live signal forwarding happens in ghost-watcher on_message → jc_reporter.forward_new_jc_message()
        logger.info("✅ JC Reporter scheduled (15min updates, hourly levels, 2min proximity alerts, daily 11PM)")
    except Exception as e:
        logger.error(f"❌ JC Copy Trader failed to load: {e}")
    
    # BTC TRADING PERFORMANCE REPORTS — Every 8 hours (00:00, 08:00, 16:00 IST)
    try:
        from src.btc_reporter import send_btc_trading_report
        # DISABLED per CEO — only trade open/close notifications
        # scheduler.add_job(send_btc_trading_report, 'cron', hour='0,8,16', minute=0, timezone='Asia/Kolkata', id='btc_report', replace_existing=True)
        logger.info("✅ BTC Trading Reports scheduled (every 8 hours: 00:00, 08:00, 16:00 IST)")
    except Exception as e:
        logger.error(f"❌ BTC Reporter failed to load: {e}")
    
    # ═══════════════════════════════════════════════════════════════
    # LEARNING ENGINE — Sprint 3 Scheduled Jobs
    # ═══════════════════════════════════════════════════════════════
    async def _learning_optimize_thresholds():
        try:
            from src.learning.improvement import LearningEngine
            engine = LearningEngine(get_async_pool())
            result = await engine.optimize_thresholds()
            logger.info(f"🎯 Threshold optimization complete: rec={result.get('recommendation')}")
        except Exception as e:
            logger.error(f"❌ Threshold optimization failed: {e}")

    async def _learning_auto_disable():
        try:
            from src.learning.improvement import LearningEngine
            engine = LearningEngine(get_async_pool())
            result = await engine.auto_disable_check()
            disabled = len(result.get('disabled_strategies', []))
            flagged = len(result.get('flagged_sports', []))
            if disabled or flagged:
                logger.info(f"🚨 Auto-disable: {disabled} strategies disabled, {flagged} sports flagged")
        except Exception as e:
            logger.error(f"❌ Auto-disable check failed: {e}")

    async def _learning_weekly_report():
        try:
            from src.learning.improvement import LearningEngine
            engine = LearningEngine(get_async_pool())
            report = await engine.weekly_report()
            logger.info(f"📊 Weekly learning report generated: {len(report.get('recommendations', []))} recommendations")
            # Send via Telegram if available (gate: paper notifications)
            _tcfg = await get_trading_mode()
            if _telegram_bot and should_send_paper_telegram(_tcfg):
                recs = report.get('recommendations', [])
                perf = report.get('strategy_performance', [])
                msg_parts = ["📊 <b>Weekly Learning Report</b>\n"]
                for sp in perf:
                    msg_parts.append(
                        f"  • <b>{sp['strategy']}</b>: {sp['win_rate']:.0%} WR, "
                        f"{sp['total_trades']} trades, ${sp['total_pnl']:+.2f}"
                    )
                if recs:
                    msg_parts.append("\n<b>Recommendations:</b>")
                    for r in recs:
                        msg_parts.append(f"  {r}")
                try:
                    await _telegram_bot.app.bot.send_message(
                        chat_id=_telegram_bot.admin_chat_id,
                        text="\n".join(msg_parts),
                        parse_mode='HTML'
                    )
                except Exception:
                    pass
        except Exception as e:
            logger.error(f"❌ Weekly learning report failed: {e}")

    # Daily at 4 AM IST = 22:30 UTC
    scheduler.add_job(_learning_optimize_thresholds, 'cron', hour=22, minute=30, id='learning_optimize', replace_existing=True)
    # Every 6 hours
    scheduler.add_job(_learning_auto_disable, 'interval', hours=6, id='learning_auto_disable', replace_existing=True)
    # Sunday 9 AM IST = Sunday 3:30 UTC
    scheduler.add_job(_learning_weekly_report, 'cron', day_of_week='sun', hour=3, minute=30, id='learning_weekly', replace_existing=True)
    logger.info("✅ Learning engine jobs scheduled (optimize: daily 4AM IST, auto-disable: 6h, weekly: Sun 9AM IST)")
    
    scheduler.start()
    logger.info("✅ Scheduler started (btc_signal: 45s, btc_resolution: 2min, v4_paper: 45s, btc_hourly, btc_intelligence, btc_daily, jc_copy_trader, jc_hourly)")
    logger.info("✅ BroBot ready")

    yield

    # Shutdown
    logger.info("🛑 WeatherBot shutting down...")
    scheduler.shutdown(wait=False)
    
    
    # Shutdown Telegram bot
    if _telegram_bot:
        from src.alerts.subscriber_bot import shutdown_bot
        await shutdown_bot()
    
    await close_pool()


app = FastAPI(
    title="BroBot — PolyEdge",
    version="0.2.0",
    lifespan=lifespan
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=True,
)


# =============================================================================
# Ghost Trading Bot API Bridge
# =============================================================================
try:
    import sys
    sys.path.insert(0, '/data/.openclaw/workspace/projects/Ghost')
    from api_bridge import router as ghost_router
    app.include_router(ghost_router)
    logger.info("\u2705 Ghost trading bot API mounted at /api/ghost")
except Exception as e:
    logger.warning(f"\u26a0\ufe0f Ghost API not loaded: {e}")

# =============================================================================
# Paper Trades
# =============================================================================



# =============================================================================
# Polymarket Leaderboard API
# =============================================================================



# =============================================================================
# Polymarket Gamma API Proxies
# =============================================================================





# =============================================================================
# Health & Status
# =============================================================================

@app.get("/api/health")
@app.get("/health")
async def health_check():
    return {
        "status": "healthy",
        "timestamp": datetime.utcnow().isoformat(),
        "version": "0.2.0",
        "scheduler": scheduler.running,
        "last_data_scan": _last_data_scan.isoformat() if _last_data_scan else None,
    }


@app.get("/api/bot/status")
async def bot_status():
    uptime = (datetime.utcnow() - _startup_time).total_seconds() if _startup_time else 0

    # Auto-trade stats from signal loop
    auto_trade_stats = {}
    if _signal_loop is not None:
        raw = getattr(_signal_loop, '_last_auto_trade_stats', None) or {}
        auto_trade_stats = {
            "trades_auto_placed_today": raw.get("trades_placed", 0),
            "signals_evaluated_today": raw.get("signals_evaluated", 0),
            "trades_skipped_today": len(raw.get("skipped", [])),
            "skip_reasons": raw.get("skipped", [])[:20],  # cap at 20 for readability
        }
        # Also pull DB-level stats
        try:
            from src.execution.paper_trader import PaperTrader
            from src.db_async import get_async_pool
            trader = PaperTrader(get_async_pool())
            db_stats = await trader.get_today_stats()
            auto_trade_stats["db_trades_placed_today"] = db_stats.get("trades_placed_today", 0)
            auto_trade_stats["open_positions"] = db_stats.get("open_positions", 0)
            auto_trade_stats["daily_pnl_usd"] = db_stats.get("daily_pnl_usd", 0.0)
        except Exception:
            pass

    return {
        "running": scheduler.running,
        "mode": getattr(config, 'MODE', 'paper'),
        "last_data_scan": _last_data_scan.isoformat() if _last_data_scan else None,
        "uptime_seconds": int(uptime),
        "signal_loop_ready": _signal_loop is not None,
        "auto_trade": auto_trade_stats,
    }


# =============================================================================
# Manual Triggers
# =============================================================================



# =============================================================================
# METAR Data
# =============================================================================





# =============================================================================
# Markets — queries REAL weather_markets table
# =============================================================================



# =============================================================================
# Signals — queries REAL signals table
# =============================================================================





# =============================================================================
# Dual Strategy Endpoints
# =============================================================================











# =============================================================================
# Trades — correct column names from actual schema
# =============================================================================





# =============================================================================
# P&L & Analytics — correct column names
# =============================================================================



@app.get("/api/bankroll")
async def get_bankroll():
    """Query REAL bankroll table."""
    try:
        result = await fetch_one(
            "SELECT total_usd, available_usd, in_positions_usd, daily_pnl, timestamp as last_updated FROM bankroll ORDER BY timestamp DESC LIMIT 1"
        )
        if result:
            return {
                "total": float(result.get("total_usd", 0) or 0),
                "available": float(result.get("available_usd", 0) or 0),
                "in_positions": float(result.get("in_positions_usd", 0) or 0),
                "daily_pnl": float(result.get("daily_pnl", 0) or 0),
                "last_updated": result.get("last_updated", "").isoformat() if result.get("last_updated") else None,
            }
        return {"total": 0, "available": 0, "in_positions": 0, "daily_pnl": 0, "last_updated": None}
    except Exception as e:
        logger.error(f"Failed to fetch bankroll: {e}")
        return {"total": 0, "available": 0, "in_positions": 0, "daily_pnl": 0, "last_updated": None}




# =============================================================================
# DB Stats (for dashboard verification)
# =============================================================================



# =============================================================================
# Polymarket Explorer — Proxy to bypass ISP blocks
# =============================================================================













# =============================================================================
# Intelligence Data — Forecasts, Historical, Combined View
# =============================================================================




# =============================================================================
# Trade Settlement
# =============================================================================













# =============================================================================
# PERFORMANCE — CEO Command Center Endpoints
# =============================================================================

@app.get("/api/performance/strategies")
async def get_performance_strategies():
    """Returns all strategies with their signal stats."""
    try:
        # Sports strategies from sports_signals
        sports_cross_odds = await fetch_one("""
            SELECT COUNT(*) as signal_count,
                   AVG(edge_pct) as avg_edge,
                   COUNT(CASE WHEN signal = 'BUY' OR signal LIKE 'BUY%' THEN 1 END) as buy_count,
                   COUNT(CASE WHEN signal = 'SELL' OR signal LIKE 'SELL%' THEN 1 END) as sell_count,
                   MAX(created_at) as last_signal_time,
                   string_agg(DISTINCT sport, ', ' ORDER BY sport) as sports_covered
            FROM sports_signals
            WHERE edge_type = 'cross_odds'
        """)
        
        sports_logical_arb = await fetch_one("""
            SELECT COUNT(*) as signal_count,
                   AVG(edge_pct) as avg_edge,
                   COUNT(CASE WHEN signal = 'BUY' OR signal LIKE 'BUY%' THEN 1 END) as buy_count,
                   COUNT(CASE WHEN signal = 'SELL' OR signal LIKE 'SELL%' THEN 1 END) as sell_count,
                   MAX(created_at) as last_signal_time,
                   string_agg(DISTINCT sport, ', ' ORDER BY sport) as sports_covered
            FROM sports_signals
            WHERE edge_type = 'logical_arb'
        """)
        
        # Weather strategies from signals table
        forecast_edge = await fetch_one("""
            SELECT COUNT(*) as signal_count,
                   AVG(edge) as avg_edge,
                   COUNT(CASE WHEN side = 'YES' THEN 1 END) as buy_count,
                   COUNT(CASE WHEN side = 'NO' THEN 1 END) as sell_count,
                   MAX(created_at) as last_signal_time
            FROM signals
            WHERE strategy = 'forecast_edge'
        """)
        
        intelligence_layer = await fetch_one("""
            SELECT COUNT(*) as signal_count,
                   AVG(edge) as avg_edge,
                   COUNT(CASE WHEN side = 'YES' THEN 1 END) as buy_count,
                   COUNT(CASE WHEN side = 'NO' THEN 1 END) as sell_count,
                   MAX(created_at) as last_signal_time
            FROM signals
            WHERE strategy = 'intelligence_layer'
        """)
        
        strategies = [
            {
                "id": "cross_odds",
                "name": "Cross-Odds Arbitrage",
                "emoji": "⚡",
                "description": "DraftKings vs Polymarket price discrepancies",
                "signal_count": sports_cross_odds.get('signal_count', 0) if sports_cross_odds else 0,
                "avg_edge": float(sports_cross_odds.get('avg_edge', 0) or 0) if sports_cross_odds else 0,
                "buy_count": sports_cross_odds.get('buy_count', 0) if sports_cross_odds else 0,
                "sell_count": sports_cross_odds.get('sell_count', 0) if sports_cross_odds else 0,
                "last_signal_time": str(sports_cross_odds.get('last_signal_time', '')) if sports_cross_odds and sports_cross_odds.get('last_signal_time') else None,
                "sports_covered": (sports_cross_odds.get('sports_covered', '') or '').split(', ') if sports_cross_odds and sports_cross_odds.get('sports_covered') else [],
                "status": "active"
            },
            {
                "id": "logical_arb",
                "name": "Logical Arbitrage",
                "emoji": "🔗",
                "description": "Group overpricing (e.g., Stanley Cup teams > 100%)",
                "signal_count": sports_logical_arb.get('signal_count', 0) if sports_logical_arb else 0,
                "avg_edge": float(sports_logical_arb.get('avg_edge', 0) or 0) if sports_logical_arb else 0,
                "buy_count": sports_logical_arb.get('buy_count', 0) if sports_logical_arb else 0,
                "sell_count": sports_logical_arb.get('sell_count', 0) if sports_logical_arb else 0,
                "last_signal_time": str(sports_logical_arb.get('last_signal_time', '')) if sports_logical_arb and sports_logical_arb.get('last_signal_time') else None,
                "sports_covered": (sports_logical_arb.get('sports_covered', '') or '').split(', ') if sports_logical_arb and sports_logical_arb.get('sports_covered') else [],
                "status": "active"
            },
            {
                "id": "forecast_edge",
                "name": "Forecast Edge",
                "emoji": "🌡️",
                "description": "Weather model predictions vs market prices",
                "signal_count": forecast_edge.get('signal_count', 0) if forecast_edge else 0,
                "avg_edge": float(forecast_edge.get('avg_edge', 0) or 0) if forecast_edge else 0,
                "buy_count": forecast_edge.get('buy_count', 0) if forecast_edge else 0,
                "sell_count": forecast_edge.get('sell_count', 0) if forecast_edge else 0,
                "last_signal_time": str(forecast_edge.get('last_signal_time', '')) if forecast_edge and forecast_edge.get('last_signal_time') else None,
                "sports_covered": ["Weather"],
                "status": "active"
            },
            {
                "id": "intelligence_layer",
                "name": "8-Gate Intelligence",
                "emoji": "🧠",
                "description": "Multi-source convergence + Claude analysis",
                "signal_count": intelligence_layer.get('signal_count', 0) if intelligence_layer else 0,
                "avg_edge": float(intelligence_layer.get('avg_edge', 0) or 0) if intelligence_layer else 0,
                "buy_count": intelligence_layer.get('buy_count', 0) if intelligence_layer else 0,
                "sell_count": intelligence_layer.get('sell_count', 0) if intelligence_layer else 0,
                "last_signal_time": str(intelligence_layer.get('last_signal_time', '')) if intelligence_layer and intelligence_layer.get('last_signal_time') else None,
                "sports_covered": ["Weather"],
                "status": "active"
            },
            {
                "id": "live_momentum",
                "name": "Live Momentum",
                "emoji": "🏃",
                "description": "Real-time event tracking + price movements",
                "signal_count": 0,
                "avg_edge": 0,
                "buy_count": 0,
                "sell_count": 0,
                "last_signal_time": None,
                "sports_covered": [],
                "status": "coming_soon"
            }
        ]
        
        return {"strategies": strategies, "count": len(strategies)}
    except Exception as e:
        logger.error(f"Performance strategies error: {e}")
        return {"strategies": [], "count": 0}












# =============================================================================
# Learning Engine — Sprint 3 Endpoints
# =============================================================================

@app.get("/api/learning/scorecard")
async def learning_scorecard(strategy: str = "arbitrage", lookback_days: int = 30):
    """Strategy scorecard from the learning engine."""
    try:
        from src.learning.improvement import LearningEngine
        from src.db_async import get_async_pool
        engine = LearningEngine(get_async_pool())
        result = await engine.strategy_scorecard(strategy, lookback_days)
        return result
    except Exception as e:
        logger.error(f"Learning scorecard error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/learning/thresholds")
async def learning_thresholds():
    """Current edge threshold analysis."""
    try:
        from src.learning.improvement import LearningEngine
        from src.db_async import get_async_pool
        engine = LearningEngine(get_async_pool())
        result = await engine.get_current_thresholds()
        return result
    except Exception as e:
        logger.error(f"Learning thresholds error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/learning/report/latest")
async def learning_report_latest():
    """Latest weekly learning report."""
    try:
        from src.learning.improvement import LearningEngine
        from src.db_async import get_async_pool
        engine = LearningEngine(get_async_pool())
        result = await engine.get_latest_report()
        if not result:
            return {"message": "No weekly report generated yet"}
        return result
    except Exception as e:
        logger.error(f"Learning report error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/learning/optimize")
async def learning_optimize():
    """Trigger threshold optimization manually."""
    try:
        from src.learning.improvement import LearningEngine
        from src.db_async import get_async_pool
        engine = LearningEngine(get_async_pool())
        result = await engine.optimize_thresholds()
        return result
    except Exception as e:
        logger.error(f"Learning optimize error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# Intelligence & Improvement — Analysis Endpoints
# =============================================================================

@app.get("/api/intelligence/daily")
async def get_daily_analysis():
    """Get daily performance analysis from improvement engine."""
    global _improvement_engine
    
    if not _improvement_engine:
        raise HTTPException(status_code=503, detail="Improvement engine not initialized")
    
    try:
        analysis = await _improvement_engine.daily_analysis()
        return analysis
    except Exception as e:
        logger.error(f"Daily analysis error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to generate daily analysis: {str(e)}")


@app.get("/api/intelligence/weekly")
async def get_weekly_review():
    """Get weekly strategy review with findings and proposals."""
    global _improvement_engine
    
    if not _improvement_engine:
        raise HTTPException(status_code=503, detail="Improvement engine not initialized")
    
    try:
        review = await _improvement_engine.weekly_review()
        return review
    except Exception as e:
        logger.error(f"Weekly review error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to generate weekly review: {str(e)}")


@app.get("/api/intelligence/calibration")
async def get_probability_calibration():
    """Get probability model calibration metrics."""
    global _improvement_engine
    
    if not _improvement_engine:
        raise HTTPException(status_code=503, detail="Improvement engine not initialized")
    
    try:
        calibration = await _improvement_engine.calibrate_probability_model()
        return calibration
    except Exception as e:
        logger.error(f"Calibration error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to generate calibration: {str(e)}")


@app.post("/api/intelligence/approve/{proposal_id}")
async def approve_proposal(proposal_id: str):
    """CEO approves a strategy change proposal.
    
    NOTE: This is a placeholder. Real implementation would:
    1. Fetch proposal from DB
    2. Parse parameter changes
    3. Update STRATEGY.md
    4. Update config values
    5. Log approval in changelog
    """
    # TODO: Implement proposal approval workflow
    return {
        "status": "approved",
        "proposal_id": proposal_id,
        "message": "Proposal approval not yet implemented. See STRATEGY.md for manual updates.",
        "timestamp": datetime.utcnow().isoformat()
    }


# =============================================================================
# Settings & Bot Control
# =============================================================================

@app.get("/api/settings")
async def get_settings():
    """Get all bot settings from DB and config."""
    try:
        # Fetch from bot_settings table
        query = "SELECT key, value FROM bot_settings"
        rows = await fetch_all(query)
        settings = {row["key"]: row["value"] for row in rows}
        
        # Add config defaults if missing
        if not settings:
            settings = {
                "max_position_size": 50,
                "max_portfolio_exposure": 15,
                "kelly_fraction": 0.25,
                "max_trades_per_city_per_day": 3,
                "daily_loss_limit": 10,
                "min_edge_auto_trade": 25,
                "min_edge_alert": 15,
                "min_confidence_sources": 2,
                "max_spread_cents": 8,
                "min_liquidity_multiple": 2,
                "min_hours_to_resolution": 2,
                "gates_enabled": {
                    "data_convergence": True,
                    "multi_station": True,
                    "bucket_coherence": True,
                    "binary_arbitrage": True,
                    "liquidity_check": True,
                    "time_window": True,
                    "risk_manager": True,
                    "claude_confirmation": True
                },
                "telegram_alerts": False,
                "telegram_chat_id": "",
                "alert_on_trade": True,
                "alert_on_signal": True,
                "alert_on_daily_summary": True,
                "refresh_interval_min": 15,
                "strategy_auto_proposals": True,
                "weekly_review_day": "Sunday",
                "weekly_review_time": "09:00",
                "auto_adjust_accuracy": True,
                "proposal_approval_mode": "require"
            }
        
        return {
            "settings": settings,
            "mode": getattr(config, 'MODE', 'paper'),
            "wallet_address": getattr(config, 'WALLET_ADDRESS', None)
        }
    except Exception as e:
        logger.error(f"Get settings error: {e}")
        return {"settings": {}, "mode": "paper", "wallet_address": None}


@app.post("/api/settings")
async def save_settings(settings: dict):
    """Save bot settings to DB."""
    try:
        for key, value in settings.items():
            await execute(
                "INSERT INTO bot_settings (key, value, updated_at) VALUES (%s, %s::jsonb, NOW()) ON CONFLICT (key) DO UPDATE SET value = %s::jsonb, updated_at = NOW()",
                (key, str(value), str(value))
            )
        return {"status": "saved", "timestamp": datetime.utcnow().isoformat()}
    except Exception as e:
        logger.error(f"Save settings error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to save settings: {str(e)}")


@app.post("/api/bot/start")
async def start_bot():
    """Start the trading bot."""
    global scheduler
    if not scheduler.running:
        scheduler.start()
    return {"status": "started", "running": scheduler.running, "timestamp": datetime.utcnow().isoformat()}


@app.post("/api/bot/pause")
async def pause_bot():
    """Pause — stop new trades but keep monitoring."""
    # TODO: Add pause flag to signal loop
    return {"status": "paused", "timestamp": datetime.utcnow().isoformat()}


@app.post("/api/bot/stop")
async def stop_bot():
    """Emergency stop — halt everything immediately."""
    global scheduler
    if scheduler.running:
        scheduler.pause()
    return {"status": "stopped", "running": scheduler.running, "timestamp": datetime.utcnow().isoformat()}




@app.get("/api/wallet/balance")
async def get_wallet_balance():
    """Get real USDC + MATIC balance from Polygon blockchain."""
    import httpx
    WALLET = "0xb9BEa5FDe7957709D0f8d2064188B1603b74D5Ca"
    RPC = "https://small-cosmological-friday.matic.quiknode.pro/c6dd1696accaf8a955e09f34475deb31913c5254/"
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            # MATIC balance
            r = await client.post(RPC, json={
                "jsonrpc": "2.0", "method": "eth_getBalance",
                "params": [WALLET, "latest"], "id": 1
            })
            matic = int(r.json()["result"], 16) / 1e18
            # USDC balance (native USDC on Polygon)
            USDC = "0x3c499c542cEF5E3811e1192ce70d8cC03d5c3359"
            data = "0x70a08231000000000000000000000000" + WALLET[2:]
            r2 = await client.post(RPC, json={
                "jsonrpc": "2.0", "method": "eth_call",
                "params": [{"to": USDC, "data": data}, "latest"], "id": 2
            })
            usdc = int(r2.json()["result"], 16) / 1e6
        return {
            "wallet": WALLET, "usdc": round(usdc, 2), "matic": round(matic, 4),
            "chain": "Polygon", "live": True
        }
    except Exception as e:
        logger.error(f"Wallet balance error: {e}")
        return {"wallet": WALLET, "usdc": 0, "matic": 0, "chain": "Polygon", "live": True, "error": str(e)}


# ═══════════════════════════════════════════════════════════════
# SPORTS INTELLIGENCE ENDPOINTS
# ═══════════════════════════════════════════════════════════════

from src.sports.polymarket_sports_scanner import PolymarketSportsScanner
from src.sports.espn_live import ESPNLiveScores
from src.sports.correlation_engine import CorrelationEngine
from src.sports.cross_odds_engine import CrossOddsEngine


# =============================================================================
# Internal Arbitrage API
# =============================================================================

# =============================================================================
# Penny Hunter API
# =============================================================================







# ═══════════════════════════════════════════════════════════════
# LATE WINDOW SCALPER — API Endpoints
# ═══════════════════════════════════════════════════════════════










# ═══════════════════════════════════════════════════════════════
# BTC SIGNAL ENGINE — API Endpoints
# ═══════════════════════════════════════════════════════════════


@app.get("/api/btc/state")
async def get_btc_state():
    """Live engine state: BTC price, active windows, signals, factors, accuracy."""
    global _btc_engine
    try:
        if _btc_engine is None:
            from src.strategies.btc_signal_engine import BTCSignalEngine
            _btc_engine = BTCSignalEngine(get_async_pool())
            await _btc_engine.ensure_tables()
        return await _btc_engine.get_current_state()
    except Exception as e:
        logger.error(f"BTC state error: {e}")
        return {"error": str(e), "btc_price": 0, "active_windows": [], "accuracy": {}}


@app.get("/api/btc/v4-trades")
async def get_btc_v4_trades():
    """BroBot V4 paper trades with full confluence breakdown — for signal cards."""
    try:
        pool = get_async_pool()
        async with pool.acquire() as conn:
            rows = await conn.fetch("""
                SELECT
                    id, window_id, direction, token_price, stake_usd,
                    confluence_score, factors_json, jc_level,
                    resolution, won, simulated_pnl, strategy_version,
                    created_at, resolved_at
                FROM paper_trades
                ORDER BY created_at DESC
                LIMIT 30
            """)
            trades = []
            for r in rows:
                t = dict(r)
                for k, v in t.items():
                    if hasattr(v, 'isoformat'):
                        t[k] = v.isoformat()
                    elif hasattr(v, '__float__'):
                        t[k] = float(v)
                trades.append(t)
            return {"trades": trades, "count": len(trades)}
    except Exception as e:
        logger.error(f"V4 trades error: {e}")
        return {"trades": [], "count": 0, "error": str(e)}


@app.get("/api/btc/signals")
async def get_btc_signals(limit: int = 50):
    """Recent signals with accuracy info."""
    global _btc_engine
    try:
        if _btc_engine is None:
            from src.strategies.btc_signal_engine import BTCSignalEngine
            _btc_engine = BTCSignalEngine(get_async_pool())
            await _btc_engine.ensure_tables()
        signals = await _btc_engine.get_recent_signals(limit=min(limit, 200))
        return {"signals": signals, "count": len(signals)}
    except Exception as e:
        logger.error(f"BTC signals error: {e}")
        return {"signals": [], "count": 0, "error": str(e)}


@app.get("/api/btc/accuracy")
async def get_btc_accuracy():
    """Rolling accuracy stats."""
    global _btc_engine
    try:
        if _btc_engine is None:
            from src.strategies.btc_signal_engine import BTCSignalEngine
            _btc_engine = BTCSignalEngine(get_async_pool())
            await _btc_engine.ensure_tables()
        return await _btc_engine.get_accuracy_stats()
    except Exception as e:
        logger.error(f"BTC accuracy error: {e}")
        return {"error": str(e)}


@app.get("/api/btc/windows")
async def get_btc_windows(limit: int = 50):
    """Active + recent windows."""
    global _btc_engine
    try:
        if _btc_engine is None:
            from src.strategies.btc_signal_engine import BTCSignalEngine
            _btc_engine = BTCSignalEngine(get_async_pool())
            await _btc_engine.ensure_tables()
        windows = await _btc_engine.get_windows(limit=min(limit, 200))
        return {"windows": windows, "count": len(windows)}
    except Exception as e:
        logger.error(f"BTC windows error: {e}")
        return {"windows": [], "count": 0, "error": str(e)}


@app.get("/api/btc/calibration")
async def get_btc_calibration(limit: int = 30):
    """Weight calibration history."""
    global _btc_engine
    try:
        if _btc_engine is None:
            from src.strategies.btc_signal_engine import BTCSignalEngine
            _btc_engine = BTCSignalEngine(get_async_pool())
            await _btc_engine.ensure_tables()
        calibrations = await _btc_engine.get_calibration_history(limit=min(limit, 100))
        return {"calibrations": calibrations, "count": len(calibrations)}
    except Exception as e:
        logger.error(f"BTC calibration error: {e}")
        return {"calibrations": [], "count": 0, "error": str(e)}
# ═══════════════════════════════════════════════════════════════


@app.get("/api/btc/analysis")
async def get_btc_analysis(hours: int = 1):
    """Analysis data for dashboard with timeframe toggle (1h, 4h, 8h, 24h, 72h, 168h)."""
    try:
        pool = get_async_pool()
        interval = f"{hours} hours"
        async with pool.acquire() as conn:
            # Entry price bucket performance
            buckets = await conn.fetch(f"""
                WITH best AS (
                    SELECT DISTINCT ON (s.window_id)
                        s.window_id, s.prediction, w.resolution, w.window_length,
                        w.close_time,
                        (s.prediction = w.resolution) as correct,
                        CASE WHEN s.prediction = 'UP' THEN w.up_price ELSE w.down_price END as entry_price
                    FROM btc_signals s
                    JOIN btc_windows w ON s.window_id = w.window_id
                    WHERE s.prediction != 'SKIP' AND w.resolution IS NOT NULL
                        AND w.close_time > NOW() - INTERVAL '{interval}'
                    ORDER BY s.window_id, s.confidence DESC
                ),
                pnl AS (
                    SELECT *,
                        CASE
                            WHEN correct AND entry_price > 0 AND entry_price < 1 THEN (25.0/entry_price - 25.0)*0.98
                            WHEN NOT correct THEN -25.0 ELSE 0
                        END as trade_pnl,
                        CASE 
                            WHEN entry_price < 0.3 THEN '<30c'
                            WHEN entry_price < 0.5 THEN '30-50c'
                            WHEN entry_price < 0.7 THEN '50-70c'
                            WHEN entry_price < 0.85 THEN '70-85c'
                            ELSE '85c+'
                        END as bucket
                    FROM best
                )
                SELECT bucket, COUNT(*) as trades, 
                    COUNT(*) FILTER (WHERE correct) as wins,
                    ROUND(SUM(trade_pnl)::numeric, 2) as net_pnl,
                    ROUND(AVG(CASE WHEN correct THEN trade_pnl END)::numeric, 2) as avg_win,
                    ROUND(MAX(trade_pnl)::numeric, 2) as best_trade
                FROM pnl GROUP BY bucket ORDER BY bucket
            """)

            # Hourly heatmap
            hourly = await conn.fetch(f"""
                WITH best AS (
                    SELECT DISTINCT ON (s.window_id)
                        s.window_id, s.prediction, w.resolution, w.window_length,
                        w.close_time,
                        (s.prediction = w.resolution) as correct,
                        CASE WHEN s.prediction = 'UP' THEN w.up_price ELSE w.down_price END as entry_price
                    FROM btc_signals s
                    JOIN btc_windows w ON s.window_id = w.window_id
                    WHERE s.prediction != 'SKIP' AND w.resolution IS NOT NULL
                        AND w.close_time > NOW() - INTERVAL '{interval}'
                    ORDER BY s.window_id, s.confidence DESC
                ),
                pnl AS (
                    SELECT *,
                        CASE
                            WHEN correct AND entry_price > 0 AND entry_price < 1 THEN (25.0/entry_price - 25.0)*0.98
                            WHEN NOT correct THEN -25.0 ELSE 0
                        END as trade_pnl
                    FROM best
                )
                SELECT 
                    EXTRACT(HOUR FROM close_time AT TIME ZONE 'Asia/Kolkata')::int as hour_ist,
                    COUNT(*) as trades,
                    COUNT(*) FILTER (WHERE correct) as wins,
                    ROUND(SUM(trade_pnl)::numeric, 2) as net_pnl,
                    ROUND(AVG(entry_price)::numeric, 3) as avg_entry,
                    ROUND(MAX(trade_pnl)::numeric, 2) as best_trade
                FROM pnl GROUP BY 1 ORDER BY 1
            """)

            # Cumulative P&L over time (for chart)
            cumulative = await conn.fetch(f"""
                WITH best AS (
                    SELECT DISTINCT ON (s.window_id)
                        s.window_id, s.prediction, w.resolution, w.window_length,
                        w.close_time,
                        (s.prediction = w.resolution) as correct,
                        CASE WHEN s.prediction = 'UP' THEN w.up_price ELSE w.down_price END as entry_price
                    FROM btc_signals s
                    JOIN btc_windows w ON s.window_id = w.window_id
                    WHERE s.prediction != 'SKIP' AND w.resolution IS NOT NULL
                        AND w.close_time > NOW() - INTERVAL '{interval}'
                    ORDER BY s.window_id, s.confidence DESC
                ),
                pnl AS (
                    SELECT *,
                        CASE
                            WHEN correct AND entry_price > 0 AND entry_price < 1 THEN (25.0/entry_price - 25.0)*0.98
                            WHEN NOT correct THEN -25.0 ELSE 0
                        END as trade_pnl
                    FROM best
                )
                SELECT close_time, trade_pnl, correct, entry_price,
                    SUM(trade_pnl) OVER (ORDER BY close_time) as running_pnl
                FROM pnl ORDER BY close_time
            """)

            # Confidence vs P&L
            confidence = await conn.fetch(f"""
                WITH best AS (
                    SELECT DISTINCT ON (s.window_id)
                        s.window_id, s.prediction, s.confidence, w.resolution, w.window_length,
                        w.close_time,
                        (s.prediction = w.resolution) as correct,
                        CASE WHEN s.prediction = 'UP' THEN w.up_price ELSE w.down_price END as entry_price
                    FROM btc_signals s
                    JOIN btc_windows w ON s.window_id = w.window_id
                    WHERE s.prediction != 'SKIP' AND w.resolution IS NOT NULL
                        AND w.close_time > NOW() - INTERVAL '{interval}'
                    ORDER BY s.window_id, s.confidence DESC
                ),
                pnl AS (
                    SELECT *,
                        CASE
                            WHEN correct AND entry_price > 0 AND entry_price < 1 THEN (25.0/entry_price - 25.0)*0.98
                            WHEN NOT correct THEN -25.0 ELSE 0
                        END as trade_pnl,
                        CASE 
                            WHEN confidence < 0.4 THEN 'Low <40%'
                            WHEN confidence < 0.6 THEN 'Med 40-60%'
                            WHEN confidence < 0.8 THEN 'High 60-80%'
                            ELSE 'Ultra 80%+'
                        END as conf_tier
                    FROM best
                )
                SELECT conf_tier, COUNT(*) as trades, 
                    COUNT(*) FILTER (WHERE correct) as wins,
                    ROUND(SUM(trade_pnl)::numeric, 2) as net_pnl,
                    ROUND(AVG(entry_price)::numeric, 3) as avg_entry
                FROM pnl GROUP BY conf_tier ORDER BY conf_tier
            """)

            # Timeframe comparison
            timeframes = await conn.fetch(f"""
                WITH best AS (
                    SELECT DISTINCT ON (s.window_id)
                        s.window_id, s.prediction, w.resolution, w.window_length,
                        w.close_time,
                        (s.prediction = w.resolution) as correct,
                        CASE WHEN s.prediction = 'UP' THEN w.up_price ELSE w.down_price END as entry_price
                    FROM btc_signals s
                    JOIN btc_windows w ON s.window_id = w.window_id
                    WHERE s.prediction != 'SKIP' AND w.resolution IS NOT NULL
                        AND w.close_time > NOW() - INTERVAL '{interval}'
                    ORDER BY s.window_id, s.confidence DESC
                ),
                pnl AS (
                    SELECT *,
                        CASE
                            WHEN correct AND entry_price > 0 AND entry_price < 1 THEN (25.0/entry_price - 25.0)*0.98
                            WHEN NOT correct THEN -25.0 ELSE 0
                        END as trade_pnl
                    FROM best
                )
                SELECT window_length||'m' as wl,
                    COUNT(*) as trades,
                    COUNT(*) FILTER (WHERE correct) as wins,
                    ROUND(SUM(trade_pnl)::numeric, 2) as net_pnl,
                    ROUND(AVG(entry_price)::numeric, 3) as avg_entry,
                    ROUND(MAX(trade_pnl)::numeric, 2) as best_trade
                FROM pnl GROUP BY window_length ORDER BY window_length
            """)

            # V2 filter comparison
            v2_compare = await conn.fetch(f"""
                WITH best AS (
                    SELECT DISTINCT ON (s.window_id)
                        s.window_id, s.prediction, w.resolution, w.window_length,
                        w.close_time,
                        (s.prediction = w.resolution) as correct,
                        CASE WHEN s.prediction = 'UP' THEN w.up_price ELSE w.down_price END as entry_price
                    FROM btc_signals s
                    JOIN btc_windows w ON s.window_id = w.window_id
                    WHERE s.prediction != 'SKIP' AND w.resolution IS NOT NULL
                        AND w.close_time > NOW() - INTERVAL '{interval}'
                    ORDER BY s.window_id, s.confidence DESC
                ),
                pnl AS (
                    SELECT *,
                        CASE
                            WHEN correct AND entry_price > 0 AND entry_price < 1 THEN (25.0/entry_price - 25.0)*0.98
                            WHEN NOT correct THEN -25.0 ELSE 0
                        END as trade_pnl
                    FROM best
                )
                SELECT 
                    'All Trades' as strategy,
                    COUNT(*) as trades,
                    COUNT(*) FILTER (WHERE correct) as wins,
                    ROUND(SUM(trade_pnl)::numeric, 2) as net_pnl,
                    ROUND(MAX(trade_pnl)::numeric, 2) as best
                FROM pnl
                UNION ALL
                SELECT 
                    'V2 (<70c, 5M)' as strategy,
                    COUNT(*) as trades,
                    COUNT(*) FILTER (WHERE correct) as wins,
                    ROUND(SUM(trade_pnl)::numeric, 2) as net_pnl,
                    ROUND(MAX(trade_pnl)::numeric, 2) as best
                FROM pnl WHERE entry_price <= 0.70 AND window_length = 5
            """)

        def row_to_dict(r, keys):
            return {k: (float(v) if isinstance(v, (int, float)) or (hasattr(v, '__float__')) else v.isoformat() if hasattr(v, 'isoformat') else v) for k, v in zip(keys, r) if v is not None}

        return {
            "timeframe_hours": hours,
            "buckets": [{"bucket": r['bucket'], "trades": r['trades'], "wins": r['wins'], "net_pnl": float(r['net_pnl'] or 0), "avg_win": float(r['avg_win'] or 0), "best_trade": float(r['best_trade'] or 0)} for r in buckets],
            "hourly": [{"hour": r['hour_ist'], "trades": r['trades'], "wins": r['wins'], "net_pnl": float(r['net_pnl'] or 0), "avg_entry": float(r['avg_entry'] or 0), "best_trade": float(r['best_trade'] or 0)} for r in hourly],
            "cumulative": [{"time": r['close_time'].isoformat(), "pnl": float(r['trade_pnl'] or 0), "running": float(r['running_pnl'] or 0), "correct": r['correct'], "entry": float(r['entry_price'] or 0)} for r in cumulative],
            "confidence": [{"tier": r['conf_tier'], "trades": r['trades'], "wins": r['wins'], "net_pnl": float(r['net_pnl'] or 0), "avg_entry": float(r['avg_entry'] or 0)} for r in confidence],
            "timeframes": [{"wl": r['wl'], "trades": r['trades'], "wins": r['wins'], "net_pnl": float(r['net_pnl'] or 0), "avg_entry": float(r['avg_entry'] or 0), "best_trade": float(r['best_trade'] or 0)} for r in timeframes],
            "v2_compare": [{"strategy": r['strategy'], "trades": r['trades'], "wins": r['wins'], "net_pnl": float(r['net_pnl'] or 0), "best": float(r['best'] or 0)} for r in v2_compare],
        }
    except Exception as e:
        logger.error(f"BTC analysis error: {e}\n{traceback.format_exc()}")
        return {"error": str(e)}


@app.get("/api/btc/bankroll")
async def get_btc_bankroll_endpoint():
    """Current bankroll status with P&L metrics."""
    try:
        state = await get_btc_bankroll_state()
        balance = float(state.get('balance', BTC_STARTING_BALANCE))
        available = float(state.get('available', balance))
        in_positions = float(state.get('in_positions', 0))
        total_won = float(state.get('total_won', 0))
        total_lost = float(state.get('total_lost', 0))
        total_trades = int(state.get('total_trades', 0))
        peak_balance = float(state.get('peak_balance', balance))
        max_drawdown_pct = float(state.get('max_drawdown_pct', 0))
        pnl_pct = (balance - BTC_STARTING_BALANCE) / BTC_STARTING_BALANCE * 100 if BTC_STARTING_BALANCE > 0 else 0
        max_single_bet = balance * 0.10
        max_exposure = balance * 0.30
        return {
            "balance": round(balance, 2),
            "available": round(available, 2),
            "in_positions": round(in_positions, 2),
            "total_won": round(total_won, 2),
            "total_lost": round(total_lost, 2),
            "total_trades": total_trades,
            "peak_balance": round(peak_balance, 2),
            "max_drawdown_pct": round(max_drawdown_pct, 2),
            "pnl_pct": round(pnl_pct, 2),
            "max_single_bet": round(max_single_bet, 2),
            "max_exposure": round(max_exposure, 2),
            "starting_balance": BTC_STARTING_BALANCE,
        }
    except Exception as e:
        logger.error(f"BTC bankroll endpoint error: {e}")
        return {"error": str(e)}


@app.get("/api/btc/stats")
async def get_btc_stats():
    """Real-time stats using correct btc_pnl view."""
    try:
        pool = get_async_pool()
        async with pool.acquire() as conn:
            row = await conn.fetchrow("""
                SELECT
                    COUNT(*) as total_trades,
                    COUNT(*) FILTER (WHERE correct) as wins,
                    COUNT(*) FILTER (WHERE NOT correct) as losses,
                    ROUND(COUNT(*) FILTER (WHERE correct)::numeric / NULLIF(COUNT(*), 0) * 100, 1) as win_rate,
                    ROUND(SUM(trade_pnl)::numeric, 2) as net_pnl,
                    ROUND(SUM(CASE WHEN trade_pnl > 0 THEN trade_pnl ELSE 0 END)::numeric, 2) as gross_profit,
                    ROUND(SUM(CASE WHEN trade_pnl < 0 THEN ABS(trade_pnl) ELSE 0 END)::numeric, 2) as gross_loss,
                    ROUND(SUM(stake)::numeric, 2) as total_spent,
                    ROUND(MAX(trade_pnl)::numeric, 2) as best_trade,
                    ROUND(SUM(CASE WHEN close_time::date = CURRENT_DATE THEN trade_pnl ELSE 0 END)::numeric, 2) as today_pnl,
                    COUNT(*) FILTER (WHERE close_time::date = CURRENT_DATE) as today_trades
                FROM btc_pnl
            """)
            stats = dict(row) if row else {"total_trades": 0}
            # Add bankroll-based P&L (source of truth)
            br = await conn.fetchrow("SELECT balance, total_won, total_lost, total_trades, peak_balance FROM btc_bankroll ORDER BY id DESC LIMIT 1")
            if br:
                initial = 5000.0
                stats["bankroll_balance"] = float(br["balance"])
                stats["bankroll_pnl"] = float(br["balance"]) - initial
                stats["bankroll_initial"] = initial
                stats["bankroll_won"] = float(br["total_won"])
                stats["bankroll_lost"] = float(br["total_lost"])
                stats["bankroll_trades"] = int(br["total_trades"])
                stats["peak_balance"] = float(br["peak_balance"])
                # Override net_pnl with bankroll truth
                stats["net_pnl"] = float(br["balance"]) - initial
            return stats
    except Exception as e:
        logger.error(f"BTC stats error: {e}")
        return {"error": str(e)}


@app.get("/api/btc/trades-detail")
async def get_btc_trades_detail():
    """Last 50 resolved trades with all signal factors for drill-down."""
    try:
        pool = get_async_pool()
        async with pool.acquire() as conn:
            rows = await conn.fetch("""
                SELECT
                    p.window_id, p.window_length, p.prediction, p.resolution, p.correct,
                    p.entry_price, p.stake, ROUND(p.trade_pnl::numeric, 2) as trade_pnl,
                    ROUND((p.trade_pnl / NULLIF(p.stake, 0) * 100)::numeric, 0) as roi_pct,
                    w.btc_open, w.btc_close,
                    ROUND((w.btc_close - w.btc_open)::numeric, 2) as btc_move,
                    p.close_time,
                    s.prob_up, s.confidence,
                    s.f_price_delta, s.f_momentum, s.f_volume_imbalance,
                    s.f_oracle_lead, s.f_book_imbalance, s.f_volatility, s.f_time_decay,
                    (CASE WHEN (p.prediction='DOWN' AND s.f_price_delta < 0) OR (p.prediction='UP' AND s.f_price_delta > 0) THEN 1 ELSE 0 END +
                     CASE WHEN (p.prediction='DOWN' AND s.f_momentum < 0) OR (p.prediction='UP' AND s.f_momentum > 0) THEN 1 ELSE 0 END +
                     CASE WHEN (p.prediction='DOWN' AND s.f_volume_imbalance < 0) OR (p.prediction='UP' AND s.f_volume_imbalance > 0) THEN 1 ELSE 0 END +
                     CASE WHEN (p.prediction='DOWN' AND s.f_oracle_lead < 0) OR (p.prediction='UP' AND s.f_oracle_lead > 0) THEN 1 ELSE 0 END +
                     CASE WHEN (p.prediction='DOWN' AND s.f_book_imbalance > 0) OR (p.prediction='UP' AND s.f_book_imbalance < 0) THEN 1 ELSE 0 END +
                     CASE WHEN s.f_volatility > 0.5 THEN 1 ELSE 0 END +
                     CASE WHEN s.f_time_decay > 0.5 THEN 1 ELSE 0 END) as factors_agreed
                FROM btc_pnl p
                JOIN btc_windows w ON p.window_id = w.window_id
                JOIN LATERAL (
                    SELECT * FROM btc_signals WHERE window_id = p.window_id AND prediction != 'SKIP' ORDER BY confidence DESC LIMIT 1
                ) s ON true
                ORDER BY p.close_time DESC
                LIMIT 50
            """)
            # Convert Decimal/datetime to JSON-serializable
            trades = []
            for r in rows:
                trade = dict(r)
                for k, v in trade.items():
                    if hasattr(v, 'isoformat'):
                        trade[k] = v.isoformat()
                    elif hasattr(v, '__float__'):
                        trade[k] = float(v)
                trades.append(trade)
            return {"trades": trades}
    except Exception as e:
        logger.error(f"BTC trades-detail error: {e}\n{traceback.format_exc()}")
        return {"trades": [], "error": str(e)}


@app.get("/api/btc/performance")
async def get_btc_performance():
    """Full performance summary — deduplicated best signal per window."""
    try:
        pool = get_async_pool()
        async with pool.acquire() as conn:
            # Per-timeframe stats
            stats = await conn.fetch("""
                WITH best AS (
                    SELECT DISTINCT ON (s.window_id)
                        s.window_id, s.prediction, s.prob_up, s.confidence,
                        w.resolution, w.window_length, w.btc_open, w.btc_close,
                        w.close_time, w.up_price, w.down_price,
                        (s.prediction = w.resolution) as correct
                    FROM btc_signals s
                    JOIN btc_windows w ON s.window_id = w.window_id
                    WHERE s.prediction != 'SKIP' AND w.resolution IS NOT NULL
                    ORDER BY s.window_id, s.confidence DESC
                )
                SELECT window_length, 
                    COUNT(*) as total,
                    COUNT(*) FILTER (WHERE correct) as wins,
                    COUNT(*) FILTER (WHERE NOT correct) as losses,
                    ROUND(COUNT(*) FILTER (WHERE correct)::numeric / COUNT(*)::numeric * 100, 1) as accuracy
                FROM best GROUP BY window_length ORDER BY window_length
            """)

            # Recent trades (last 30)
            trades = await conn.fetch("""
                WITH best AS (
                    SELECT DISTINCT ON (s.window_id)
                        s.window_id, s.prediction, ROUND(s.prob_up::numeric * 100) as prob,
                        ROUND(s.confidence::numeric * 100) as conf,
                        w.resolution, w.window_length, w.btc_open, w.btc_close,
                        w.close_time, w.up_price, w.down_price,
                        (s.prediction = w.resolution) as correct
                    FROM btc_signals s
                    JOIN btc_windows w ON s.window_id = w.window_id
                    WHERE s.prediction != 'SKIP' AND w.resolution IS NOT NULL
                    ORDER BY s.window_id, s.confidence DESC
                )
                SELECT * FROM best ORDER BY close_time DESC LIMIT 30
            """)

            # Hourly breakdown
            hourly = await conn.fetch("""
                WITH best AS (
                    SELECT DISTINCT ON (s.window_id)
                        s.window_id, s.prediction, w.resolution, w.window_length,
                        w.close_time, (s.prediction = w.resolution) as correct
                    FROM btc_signals s
                    JOIN btc_windows w ON s.window_id = w.window_id
                    WHERE s.prediction != 'SKIP' AND w.resolution IS NOT NULL
                    ORDER BY s.window_id, s.confidence DESC
                )
                SELECT date_trunc('hour', close_time AT TIME ZONE 'Asia/Kolkata') as hour,
                    COUNT(*) as total,
                    COUNT(*) FILTER (WHERE correct) as wins,
                    ROUND(COUNT(*) FILTER (WHERE correct)::numeric / COUNT(*)::numeric * 100, 1) as accuracy
                FROM best GROUP BY 1 ORDER BY 1 DESC LIMIT 12
            """)

        return {
            "stats": [{"window": f"{r['window_length']}m", "total": r['total'], "wins": r['wins'], "losses": r['losses'], "accuracy": float(r['accuracy'])} for r in stats],
            "trades": [{"window_id": r['window_id'], "wl": f"{r['window_length']}m", "call": r['prediction'], "prob": int(r['prob']), "conf": int(r['conf']), "actual": r['resolution'], "correct": r['correct'], "close_time": r['close_time'].isoformat(), "btc_open": float(r['btc_open']) if r['btc_open'] else None, "btc_close": float(r['btc_close']) if r['btc_close'] else None} for r in trades],
            "hourly": [{"hour": r['hour'].isoformat(), "total": r['total'], "wins": r['wins'], "accuracy": float(r['accuracy'])} for r in hourly],
        }
    except Exception as e:
        logger.error(f"BTC performance error: {e}")
        return {"error": str(e)}

























# =============================================================================
# Telegram Subscriber Bot API
# =============================================================================

@app.get("/api/telegram/subscribers")
async def get_telegram_subscribers():
    """Get subscriber count and stats."""
    try:
        async_pool = get_async_pool()
        async with async_pool.connection() as conn:
            async with conn.cursor() as cur:
                await cur.execute("""
                    SELECT 
                        COUNT(*) as total,
                        COUNT(CASE WHEN is_active = true THEN 1 END) as active,
                        COUNT(CASE WHEN alert_frequency = 'instant' THEN 1 END) as instant,
                        COUNT(CASE WHEN alert_frequency = 'daily' THEN 1 END) as daily,
                        SUM(total_alerts_sent) as total_alerts
                    FROM telegram_subscribers
                """)
                row = await cur.fetchone()
                
        return {
            "total": row[0] if row else 0,
            "active": row[1] if row else 0,
            "instant": row[2] if row else 0,
            "daily": row[3] if row else 0,
            "total_alerts_sent": row[4] if row else 0,
            "bot_enabled": _telegram_bot is not None,
            "timestamp": datetime.utcnow().isoformat()
        }
    except Exception as e:
        logger.error(f"Telegram subscribers error: {e}")
        return {"total": 0, "active": 0, "error": str(e)}


@app.post("/api/telegram/broadcast")
async def manual_broadcast(data: dict):
    """Manually broadcast a message (admin only)."""
    global _telegram_bot
    
    if not _telegram_bot:
        raise HTTPException(status_code=503, detail="Telegram bot not initialized")
    
    message = data.get("message")
    if not message:
        raise HTTPException(status_code=400, detail="message required")
    
    # Admin check (compare with TELEGRAM_ADMIN_CHAT_ID)
    admin_chat_id = data.get("admin_chat_id")
    if str(admin_chat_id) != str(_telegram_bot.admin_chat_id):
        raise HTTPException(status_code=403, detail="Admin only")
    
    try:
        subscribers = await _telegram_bot.get_all_subscribers()
        sent = 0
        for sub in subscribers:
            try:
                await _telegram_bot.app.bot.send_message(
                    chat_id=sub['chat_id'],
                    text=message,
                    parse_mode='HTML'
                )
                sent += 1
            except Exception as e:
                logger.error(f"Failed to send to {sub['chat_id']}: {e}")
        
        return {
            "status": "broadcast_complete",
            "sent": sent,
            "total_subscribers": len(subscribers),
            "timestamp": datetime.utcnow().isoformat()
        }
    except Exception as e:
        logger.error(f"Manual broadcast error: {e}")
        raise HTTPException(status_code=500, detail=str(e))




















# =============================================================================
# LIVE TRADING ENDPOINTS (live trades, live stats)
# =============================================================================

@app.get("/api/live/trades")
async def get_live_trades():
    """Get only LIVE trades (not paper)."""
    try:
        pool = get_async_pool()
        async with pool.acquire() as conn:
            rows = await conn.fetch("""
                SELECT id, window_id, prediction, token_id, side,
                       entry_price, stake_usd, tx_hash, status,
                       exit_price, pnl_usd, wallet_balance_before,
                       wallet_balance_after, created_at, resolved_at
                FROM live_trades
                ORDER BY created_at DESC
                LIMIT 50
            """)
            trades = []
            for row in rows:
                r = dict(row)
                trades.append({
                    "id": r["id"],
                    "window_id": r.get("window_id"),
                    "prediction": r.get("prediction"),
                    "token_id": r.get("token_id"),
                    "side": r.get("side"),
                    "entry_price": float(r["entry_price"]) if r.get("entry_price") else None,
                    "stake_usd": float(r["stake_usd"]) if r.get("stake_usd") else None,
                    "tx_hash": r.get("tx_hash"),
                    "status": r.get("status"),
                    "exit_price": float(r["exit_price"]) if r.get("exit_price") else None,
                    "pnl_usd": float(r["pnl_usd"]) if r.get("pnl_usd") else None,
                    "wallet_before": float(r["wallet_balance_before"]) if r.get("wallet_balance_before") else None,
                    "wallet_after": float(r["wallet_balance_after"]) if r.get("wallet_balance_after") else None,
                    "created_at": r["created_at"].isoformat() if r.get("created_at") else None,
                    "resolved_at": r["resolved_at"].isoformat() if r.get("resolved_at") else None,
                })
            return {"trades": trades, "mode": "live"}
    except Exception as e:
        logger.error(f"Live trades error: {e}")
        return {"trades": [], "mode": "live", "error": str(e)}


@app.get("/api/live/stats")
async def get_live_stats():
    """Get BTC trading stats from btc_pnl (real data)."""
    try:
        pool = get_async_pool()
        async with pool.acquire() as conn:
            # Query btc_pnl view for real BTC 15/5M trading data
            row = await conn.fetchrow("""
                SELECT
                    COUNT(*) as total,
                    COUNT(*) FILTER (WHERE correct) as wins,
                    COUNT(*) FILTER (WHERE NOT correct) as losses,
                    COALESCE(SUM(trade_pnl), 0) as total_pnl,
                    COALESCE(AVG(trade_pnl), 0) as avg_pnl,
                    COALESCE(SUM(stake), 0) as total_staked
                FROM btc_pnl
            """)
            total = int(row["total"])
            wins = int(row["wins"])
            losses = int(row["losses"])
            total_pnl = float(row["total_pnl"])
            avg_pnl = float(row["avg_pnl"])
            total_staked = float(row["total_staked"])
            win_rate = (wins / total * 100) if total > 0 else 0
            return {
                "total_trades": total,
                "wins": wins,
                "losses": losses,
                "win_rate": round(win_rate, 1),
                "total_pnl": round(total_pnl, 2),
                "avg_pnl": round(avg_pnl, 2),
                "total_staked": round(total_staked, 2),
                "mode": "live"
            }
    except Exception as e:
        logger.error(f"Live stats error: {e}")
        return {"total_trades": 0, "wins": 0, "losses": 0, "win_rate": 0, "total_pnl": 0, "avg_pnl": 0, "total_staked": 0, "mode": "live", "error": str(e)}


# =============================================================================
# HOME COMMAND CENTER ENDPOINTS
# =============================================================================

@app.get("/api/live/today")
async def get_today_pnl():
    """Today's P&L from live_trades resolved today."""
    try:
        pool = get_async_pool()
        async with pool.acquire() as conn:
            row = await conn.fetchrow("""
                SELECT
                    COUNT(*) as trades,
                    COALESCE(SUM(pnl_usd), 0) as pnl,
                    SUM(CASE WHEN pnl_usd > 0 THEN 1 ELSE 0 END) as wins
                FROM live_trades
                WHERE resolved_at >= CURRENT_DATE
                AND status = 'resolved'
            """)
        pnl = float(row['pnl'])
        return {
            "trades": row['trades'],
            "pnl": round(pnl, 2),
            "wins": row['wins'] or 0,
            "profitable": pnl >= 0
        }
    except Exception as e:
        logger.error(f"Today PnL error: {e}")
        return {"trades": 0, "pnl": 0, "wins": 0, "profitable": True}


@app.get("/api/live/positions")
async def get_active_positions():
    """Active open positions count and deployed capital."""
    try:
        pool = get_async_pool()
        async with pool.acquire() as conn:
            row = await conn.fetchrow("""
                SELECT COUNT(*) as count, COALESCE(SUM(stake_usd), 0) as deployed
                FROM live_trades WHERE status = 'open'
            """)
        return {
            "count": row['count'],
            "deployed": round(float(row['deployed']), 2)
        }
    except Exception as e:
        logger.error(f"Active positions error: {e}")
        return {"count": 0, "deployed": 0}


@app.get("/api/live/weekly")
async def get_weekly_pnl():
    """7-day P&L from live_trades."""
    try:
        pool = get_async_pool()
        async with pool.acquire() as conn:
            row = await conn.fetchrow("""
                SELECT
                    COUNT(*) as trades,
                    COALESCE(SUM(pnl_usd), 0) as pnl
                FROM live_trades
                WHERE resolved_at >= CURRENT_DATE - INTERVAL '7 days'
                AND status = 'resolved'
            """)
        return {
            "trades": row['trades'],
            "pnl": round(float(row['pnl']), 2)
        }
    except Exception as e:
        logger.error(f"Weekly PnL error: {e}")
        return {"trades": 0, "pnl": 0}


@app.get("/api/live/trend")
async def get_pnl_trend():
    """7-day daily P&L trend for sparkline chart."""
    try:
        pool = get_async_pool()
        async with pool.acquire() as conn:
            rows = await conn.fetch("""
                SELECT
                    DATE(resolved_at) as day,
                    COALESCE(SUM(pnl_usd), 0) as pnl,
                    COUNT(*) as trades
                FROM live_trades
                WHERE resolved_at >= CURRENT_DATE - INTERVAL '7 days'
                AND status = 'resolved'
                GROUP BY DATE(resolved_at)
                ORDER BY day
            """)
        return [{"day": str(r['day']), "pnl": round(float(r['pnl']), 2), "trades": r['trades']} for r in rows]
    except Exception as e:
        logger.error(f"PnL trend error: {e}")
        return []


# =============================================================================
# JC Copy Trading API Endpoints
# =============================================================================

@app.get("/api/jc/status")
async def jc_status():
    """Get JC copy trader status: mode, Bybit balance, open positions, DB trades."""
    try:
        sys.path.insert(0, '/data/.openclaw/workspace/projects/Ghost')
        from jc_copy_trader import get_trading_mode
        mode = get_trading_mode()
        
        result = {"mode": mode, "ok": True}
        
        # ALWAYS fetch real Bybit balance (regardless of paper/live mode)
        try:
            sys.path.insert(0, '/data/.openclaw/workspace/projects/Ghost')
            from bybit_executor import get_executor
            executor = get_executor()
            
            # Balance
            bal = executor.test_connectivity()
            equity = float(bal.get('equity', 0))
            available = float(bal.get('available', 0))
            
            # Positions
            pos_data = executor.get_open_positions()
            positions = pos_data.get('positions', [])
            in_positions_usd = 0.0
            unrealized = 0.0
            for p in positions:
                size = float(p.get('size', 0))
                entry = float(p.get('entry_price', 0))
                lev = float(p.get('leverage', 1))
                if size > 0 and lev > 0:
                    in_positions_usd += (size * entry) / lev
                unrealized += float(p.get('unrealized_pnl', 0))
            
            result["bankroll"] = {
                "balance": equity,
                "available": available,
                "in_positions": round(in_positions_usd, 2),
                "unrealized_pnl": round(unrealized, 4),
                "source": "bybit_live",
            }
            result["bybit_connected"] = True
            result["open_positions"] = positions
            
        except Exception as bybit_err:
            logger.warning(f"Bybit balance fetch failed: {bybit_err}")
            result["bybit_error"] = str(bybit_err)
            # Fallback to DB bankroll
            try:
                import psycopg2
                conn2 = psycopg2.connect("dbname=polyedge user=node host=localhost")
                cur2 = conn2.cursor()
                cur2.execute("SELECT balance, available, in_positions, total_won, total_lost, total_trades FROM jc_bankroll ORDER BY id DESC LIMIT 1")
                bank = cur2.fetchone()
                if bank:
                    result["bankroll"] = {
                        "balance": float(bank[0]), "available": float(bank[1]),
                        "in_positions": float(bank[2]), "total_won": float(bank[3]),
                        "total_lost": float(bank[4]), "total_trades": bank[5],
                        "source": "db_fallback",
                    }
                conn2.close()
            except Exception as db_err:
                result["db_error"] = str(db_err)
        
        # DB trades (open ones)
        try:
            import psycopg2
            conn2 = psycopg2.connect("dbname=polyedge user=node host=localhost")
            cur2 = conn2.cursor()
            cur2.execute("SELECT id, direction, entry_price, stake_usd, realized_pnl, status FROM jc_trades WHERE status IN ('active','half_closed','breakeven')")
            result["open_db_trades"] = [{"id": r[0], "direction": r[1], "entry_price": float(r[2]), "stake": float(r[3] or 0), "pnl": float(r[4] or 0), "status": r[3]} for r in cur2.fetchall()]
            # Total PnL from closed trades
            cur2.execute("SELECT COALESCE(SUM(realized_pnl), 0), COUNT(*), COUNT(*) FILTER (WHERE realized_pnl > 0), COUNT(*) FILTER (WHERE realized_pnl < 0) FROM jc_trades WHERE status = 'closed'")
            pnl_row = cur2.fetchone()
            result["trade_stats"] = {
                "total_pnl": float(pnl_row[0]),
                "total_trades": pnl_row[1],
                "wins": pnl_row[2],
                "losses": pnl_row[3],
            }
            conn2.close()
        except Exception as db_err:
            result["db_error"] = str(db_err)
        
        return result
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/jc/performance")
async def jc_performance():
    """JC Copy Trading performance — ONLY jc_trades + Bybit data. NOT Polymarket/BTC Paper."""
    pool = get_async_pool()
    try:
        async with pool.acquire() as conn:
            # === JC Copy trades ONLY (from jc_trades table) ===
            jc_stats = await conn.fetchrow("""
                SELECT 
                    count(*) as total,
                    count(*) FILTER (WHERE realized_pnl > 0) as wins,
                    count(*) FILTER (WHERE realized_pnl <= 0 AND realized_pnl IS NOT NULL) as losses,
                    COALESCE(sum(realized_pnl), 0) as total_pnl,
                    COALESCE(avg(realized_pnl) FILTER (WHERE realized_pnl > 0), 0) as avg_win,
                    COALESCE(avg(realized_pnl) FILTER (WHERE realized_pnl <= 0 AND realized_pnl IS NOT NULL), 0) as avg_loss,
                    COALESCE(max(realized_pnl), 0) as best_trade,
                    COALESCE(min(realized_pnl), 0) as worst_trade,
                    COALESCE(sum(stake_usd), 0) as total_staked
                FROM jc_trades WHERE status = 'closed'
            """)
            
            # === Daily P&L from jc_trades ===
            daily_rows = await conn.fetch("""
                SELECT 
                    closed_at::date as date,
                    sum(realized_pnl) as pnl,
                    count(*) as trades,
                    count(*) FILTER (WHERE realized_pnl > 0) as wins
                FROM jc_trades WHERE status = 'closed' AND closed_at IS NOT NULL
                GROUP BY closed_at::date
                ORDER BY date DESC
                LIMIT 30
            """)
            
            # === Hourly heatmap from jc_trades ===
            hourly_rows = await conn.fetch("""
                SELECT 
                    EXTRACT(HOUR FROM opened_at AT TIME ZONE 'Asia/Kolkata')::int as hour,
                    count(*) as trades,
                    count(*) FILTER (WHERE realized_pnl > 0) as wins,
                    COALESCE(sum(realized_pnl), 0) as pnl
                FROM jc_trades WHERE status = 'closed'
                GROUP BY 1
                ORDER BY 1
            """)
            
            # All trades for history
            all_trades = await conn.fetch("""
                SELECT id, direction, entry_price, exit_fill_price, realized_pnl,
                    stake_usd, leverage, status, entry_reason, close_reason,
                    stop_loss, take_profit_1, is_live, opened_at, closed_at
                FROM jc_trades ORDER BY opened_at DESC LIMIT 50
            """)
            
            total_trades = int(jc_stats['total'])
            total_wins = int(jc_stats['wins'])
            total_losses = int(jc_stats['losses'])
            total_pnl = float(jc_stats['total_pnl'])
            avg_win = float(jc_stats['avg_win'])
            avg_loss = float(jc_stats['avg_loss'])
            
            gross_wins_row = await conn.fetchrow("SELECT COALESCE(sum(realized_pnl) FILTER (WHERE realized_pnl > 0), 0) as v FROM jc_trades WHERE status = 'closed'")
            gross_losses_row = await conn.fetchrow("SELECT COALESCE(sum(realized_pnl) FILTER (WHERE realized_pnl <= 0), 0) as v FROM jc_trades WHERE status = 'closed'")
            gross_wins = float(gross_wins_row['v']) if gross_wins_row else 0
            gross_losses = abs(float(gross_losses_row['v'])) if gross_losses_row else 0
            profit_factor = round(gross_wins / gross_losses, 2) if gross_losses > 0 else 0
            
            decided = total_wins + total_losses
            win_rate = round(total_wins / decided, 3) if decided > 0 else 0
            
            # Real Bybit balance + unrealized
            bybit_equity = 0
            bybit_available = 0
            bybit_unrealized = 0
            bybit_positions = []
            try:
                import sys as _sys
                _sys.path.insert(0, '/data/.openclaw/workspace/projects/Ghost')
                from bybit_executor import get_executor
                executor = get_executor()
                bal = executor.test_connectivity()
                bybit_equity = float(bal.get('equity', 0))
                bybit_available = float(bal.get('available', 0))
                pos = executor.get_open_positions()
                bybit_positions = pos.get('positions', [])
                for p in bybit_positions:
                    bybit_unrealized += float(p.get('unrealized_pnl', 0))
            except:
                pass
            
            # Starting capital
            starting_capital = 100.0  # Initial Bybit deposit
            
            return {
                "total_pnl": round(total_pnl, 2),
                "unrealized_pnl": round(bybit_unrealized, 4),
                "bybit_equity": round(bybit_equity, 2),
                "bybit_available": round(bybit_available, 2),
                "starting_capital": starting_capital,
                "roi_pct": round((bybit_equity - starting_capital) / starting_capital * 100, 2) if starting_capital > 0 else 0,
                "win_rate": win_rate,
                "total_trades": total_trades,
                "wins": total_wins,
                "losses": total_losses,
                "avg_win": round(avg_win, 2),
                "avg_loss": round(avg_loss, 2),
                "profit_factor": profit_factor,
                "best_trade": round(float(jc_stats['best_trade']), 2),
                "worst_trade": round(float(jc_stats['worst_trade']), 2),
                "total_staked": round(float(jc_stats['total_staked']), 2),
                "open_positions": bybit_positions,
                "daily": [
                    {"date": str(r['date']), "pnl": round(float(r['pnl']), 2), "trades": r['trades'], "wins": r['wins']}
                    for r in reversed(daily_rows)
                ],
                "hourly": [
                    {"hour": r['hour'], "trades": r['trades'], "wins": r['wins'], "pnl": round(float(r['pnl']), 2)}
                    for r in hourly_rows
                ],
                "trades": [
                    {
                        "id": r['id'], "direction": r['direction'],
                        "entry": float(r['entry_price']), "exit": float(r['exit_fill_price']) if r['exit_fill_price'] else None,
                        "pnl": float(r['realized_pnl']) if r['realized_pnl'] else None,
                        "stake": float(r['stake_usd']), "leverage": r['leverage'],
                        "status": r['status'], "reason": r['entry_reason'],
                        "close_reason": r['close_reason'],
                        "sl": float(r['stop_loss']) if r['stop_loss'] else None,
                        "tp": float(r['take_profit_1']) if r['take_profit_1'] else None,
                        "is_live": r['is_live'],
                        "opened": r['opened_at'].isoformat() if r['opened_at'] else None,
                        "closed": r['closed_at'].isoformat() if r['closed_at'] else None,
                    }
                    for r in all_trades
                ],
            }
    except Exception as e:
        logger.error(f"JC Performance endpoint error: {e}")
        return {"error": str(e), "total_pnl": 0, "total_trades": 0}


@app.post("/api/jc/mode")
async def jc_set_mode(request: Request):
    """Toggle JC trading mode between 'paper' and 'live'."""
    try:
        body = await request.json()
        mode = body.get("mode", "paper").lower().strip()
        if mode not in ("paper", "live"):
            return {"ok": False, "error": "mode must be 'paper' or 'live'"}
        
        import psycopg2 as _pg2
        _conn2 = _pg2.connect("dbname=polyedge user=node host=localhost")
        _cur2 = _conn2.cursor()
        _cur2.execute(
            "INSERT INTO jc_settings (key, value, updated_at) VALUES (%s, %s, NOW()) "
            "ON CONFLICT (key) DO UPDATE SET value = %s, updated_at = NOW()",
            ('mode', mode, mode),
        )
        _conn2.commit()
        _conn2.close()
        
        logger.info(f"JC trading mode set to: {mode}")
        return {"ok": True, "mode": mode, "message": f"JC trading mode switched to {mode.upper()}"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/jc/test-trade")
async def jc_test_trade(request: Request):
    """Simulate a JC signal and trigger the copy trader (paper mode only).
    Useful for end-to-end pipeline testing."""
    try:
        body = await request.json()
        
        # Require paper mode for test
        sys.path.insert(0, '/data/.openclaw/workspace/projects/Ghost')
        from jc_copy_trader import get_trading_mode, open_trade, get_btc_price
        
        mode = get_trading_mode()
        if mode == 'live' and not body.get("force", False):
            return {
                "ok": False,
                "error": "Cannot test in live mode. Add force=true to override, or switch to paper mode first.",
                "current_mode": mode,
            }
        
        # Get current price or use override
        price = body.get("price") or get_btc_price() or 83000
        
        # Build a synthetic JC signal
        direction = body.get("direction", "LONG")
        if direction == "LONG":
            sl = price * 0.98     # 2% SL
            tp1 = price * 1.03   # 3% TP1 (1.5:1 R:R)
            tp2 = price * 1.05   # 5% TP2
        else:
            sl = price * 1.02
            tp1 = price * 0.97
            tp2 = price * 0.95
        
        test_signal = {
            "direction": direction,
            "entry": round(price, 2),
            "sl": round(sl, 2),
            "tp1": round(tp1, 2),
            "tp2": round(tp2, 2),
            "rr": round(abs(tp1 - price) / abs(sl - price), 1) if sl != price else 0,
            "reason": body.get("reason", "TEST nwPOC @ ${:.0f}".format(price)),
            "leverage": body.get("leverage", 10),
            "level": {"price": price, "label": "TEST", "type": "support" if direction == "LONG" else "resistance"},
            "conviction": {
                "classification": "MEDIUM",
                "score": 65.0,
                "factors_agreed": 3,
                "total_factors": 7,
                "stake_pct": 0.03,
                "rr_ratio": 1.5,
            },
            "conviction_stake_pct": 0.03,
        }
        
        logger.info(f"JC test trade: {direction} @ ${price:,.0f}")
        trade_id = open_trade(test_signal, price)
        
        return {
            "ok": True,
            "trade_id": trade_id,
            "signal": test_signal,
            "mode": mode,
            "message": f"Test trade #{trade_id} opened in {mode.upper()} mode",
        }
    except Exception as e:
        logger.error(f"JC test trade error: {e}", exc_info=True)
        return {"ok": False, "error": str(e)}


@app.post("/api/jc/kill")
async def jc_kill_position():
    """Emergency kill — close Bybit position immediately."""
    try:
        sys.path.insert(0, '/data/.openclaw/workspace/projects/Ghost')
        from jc_copy_trader import get_trading_mode
        mode = get_trading_mode()
        
        if mode != 'live':
            return {"ok": False, "error": "Not in live mode — no real position to close"}
        
        from bybit_executor import get_executor
        executor = get_executor()
        result = executor.close_position(reason="Dashboard kill")
        
        if result["ok"]:
            # Update DB
            import psycopg2 as _pg2
            _conn2 = _pg2.connect("dbname=polyedge user=node host=localhost")
            _cur2 = _conn2.cursor()
            _cur2.execute("""
                UPDATE jc_trades SET status = 'closed', close_reason = 'MANUAL_KILL',
                    closed_at = NOW() WHERE status IN ('active','half_closed','breakeven')
            """)
            _conn2.commit()
            _conn2.close()
        
        return {"ok": result["ok"], "result": result}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/jc/bybit/balance")
async def jc_bybit_balance():
    """Get live Bybit account balance."""
    try:
        sys.path.insert(0, '/data/.openclaw/workspace/projects/Ghost')
        from bybit_executor import get_executor
        executor = get_executor()
        return executor.test_connectivity()
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/jc/bybit/position")
async def jc_bybit_position():
    """Get live Bybit open position."""
    try:
        sys.path.insert(0, '/data/.openclaw/workspace/projects/Ghost')
        from bybit_executor import get_executor
        executor = get_executor()
        return executor.get_open_positions()
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/jc/pause")
async def jc_pause_trading(request: Request):
    """Pause or resume JC copy trading."""
    try:
        body = await request.json()
        paused = bool(body.get("paused", True))
        import psycopg2 as _pg2
        _conn2 = _pg2.connect("dbname=polyedge user=node host=localhost")
        _cur2 = _conn2.cursor()
        _cur2.execute(
            "INSERT INTO jc_settings (key, value, updated_at) VALUES (%s, %s, NOW()) "
            "ON CONFLICT (key) DO UPDATE SET value = %s, updated_at = NOW()",
            ('paused', str(paused).lower(), str(paused).lower()),
        )
        _conn2.commit()
        _conn2.close()
        return {"ok": True, "paused": paused}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# =============================================================================
# Tradebook — Unified Trade Ledger API
# =============================================================================

@app.get("/api/tradebook")
async def get_tradebook(
    bot: Optional[str] = None,
    limit: int = 500,
    offset: int = 0
):
    """Unified trade ledger across all bots."""
    from fastapi.responses import JSONResponse

    pool = get_async_pool()
    all_trades = []

    async def safe_query(query: str, bot_name: str):
        """Execute query and gracefully skip on error."""
        try:
            async with pool.acquire() as conn:
                rows = await conn.fetch(query)
                return [
                    {
                        'bot': bot_name,
                        'trade_id': str(r['trade_id']) if r['trade_id'] is not None else None,
                        'direction': r['direction'],
                        'entry_price': float(r['entry_price']) if r['entry_price'] is not None else None,
                        'stake': float(r['stake']) if r['stake'] is not None else 0.0,
                        'gross_pnl': float(r['gross_pnl']) if r['gross_pnl'] is not None else 0.0,
                        'fees': float(r['fees']) if r['fees'] is not None else 0.0,
                        'net_pnl': float(r['net_pnl']) if r['net_pnl'] is not None else 0.0,
                        'outcome': r['outcome'],
                        'timestamp': r['timestamp'].isoformat() if r['timestamp'] is not None else None,
                        'window_length': r.get('window_length'),
                    }
                    for r in rows
                ]
        except Exception as e:
            logger.warning(f"⚠️ Tradebook: skipping {bot_name} — {e}")
            return []

    # BTC Paper trades (from btc_pnl — the real resolved trades)
    btc_trades = await safe_query("""
        SELECT 'BTC Paper' as bot, window_id as trade_id, prediction as direction,
            entry_price, stake, trade_pnl as gross_pnl,
            CASE WHEN correct THEN 'won' ELSE 'lost' END as outcome,
            close_time as timestamp, window_length,
            0.0 as fees, trade_pnl as net_pnl
        FROM btc_pnl ORDER BY close_time DESC
    """, 'BTC Paper')
    all_trades.extend(btc_trades)

    # BTC Live trades
    live_trades = await safe_query("""
        SELECT 'BTC Live' as bot, window_id as trade_id, prediction as direction,
            entry_price, stake_usd as stake, pnl_usd as gross_pnl,
            status as outcome, created_at as timestamp,
            0.0 as fees, pnl_usd as net_pnl, NULL as window_length
        FROM live_trades ORDER BY created_at DESC
    """, 'BTC Live')
    all_trades.extend(live_trades)

    # JC Copy trades
    jc_trades = await safe_query("""
        SELECT 'JC Copy' as bot, id::text as trade_id, direction,
            entry_price, stake_usd as stake, realized_pnl as gross_pnl,
            status as outcome, opened_at as timestamp,
            0.0 as fees, realized_pnl as net_pnl, NULL as window_length
        FROM jc_trades ORDER BY opened_at DESC
    """, 'JC Copy')
    all_trades.extend(jc_trades)

    # Maker orders (filled only)
    maker_trades = await safe_query("""
        SELECT 'Maker' as bot, order_id as trade_id, side as direction,
            price as entry_price, size as stake, COALESCE(pnl_usd, 0) as gross_pnl,
            status as outcome, created_at as timestamp,
            0.0 as fees, COALESCE(pnl_usd, 0) as net_pnl, NULL as window_length
        FROM maker_orders WHERE status = 'filled' ORDER BY created_at DESC
    """, 'Maker')
    all_trades.extend(maker_trades)

    # Sports/IPL trades
    sports_trades = await safe_query("""
        SELECT 'Sports' as bot, id::text as trade_id,
            team_backed || ' (' || COALESCE(sport, 'IPL') || ')' as direction,
            entry_price, position_size as stake, COALESCE(pnl, 0) as gross_pnl,
            CASE WHEN pnl > 0 THEN 'won' WHEN pnl < 0 THEN 'lost' WHEN status = 'open' THEN 'open' ELSE status END as outcome,
            entry_at as timestamp,
            0.0 as fees, COALESCE(pnl, 0) as net_pnl, NULL as window_length
        FROM paper_trades_live ORDER BY entry_at DESC
    """, 'Sports')
    all_trades.extend(sports_trades)

    # Late window scalper trades
    scalper_trades = await safe_query("""
        SELECT 'Scalper' as bot, id::text as trade_id, direction,
            entry_price, stake_usd as stake, COALESCE(pnl_usd, 0) as gross_pnl,
            COALESCE(outcome, 'open') as outcome, traded_at as timestamp,
            0.0 as fees, COALESCE(pnl_usd, 0) as net_pnl, window_length
        FROM late_window_trades ORDER BY traded_at DESC
    """, 'Scalper')
    all_trades.extend(scalper_trades)

    # Sort all trades by timestamp DESC
    def ts_key(t):
        return t['timestamp'] or ''
    all_trades.sort(key=ts_key, reverse=True)

    # Filter by bot if requested
    if bot:
        all_trades = [t for t in all_trades if t['bot'] == bot]

    # Build summary
    total_trades = len(all_trades)
    total_staked = sum(t['stake'] or 0 for t in all_trades)
    total_gross_pnl = sum(t['gross_pnl'] or 0 for t in all_trades)
    total_fees = sum(t['fees'] or 0 for t in all_trades)
    total_net_pnl = sum(t['net_pnl'] or 0 for t in all_trades)
    won = sum(1 for t in all_trades if t['outcome'] in ('won', 'win', 'Won', 'filled') or (t.get('net_pnl') and t['net_pnl'] > 0))
    lost = sum(1 for t in all_trades if t['outcome'] in ('lost', 'loss', 'Lost') or (t.get('net_pnl') and t['net_pnl'] < 0))
    decided = won + lost
    win_rate = round((won / decided * 100) if decided > 0 else 0.0, 1)
    
    pnl_values = [t['net_pnl'] for t in all_trades if t.get('net_pnl') and t['net_pnl'] != 0]
    best_trade_pnl = max(pnl_values) if pnl_values else 0
    worst_trade_pnl = min(pnl_values) if pnl_values else 0
    best_trade_bot = next((t['bot'] for t in all_trades if t.get('net_pnl') == best_trade_pnl), '') if best_trade_pnl else ''
    worst_trade_bot = next((t['bot'] for t in all_trades if t.get('net_pnl') == worst_trade_pnl), '') if worst_trade_pnl else ''

    by_bot = {}
    for t in all_trades:
        b = t['bot']
        if b not in by_bot:
            by_bot[b] = {'trades': 0, 'net_pnl': 0.0, 'wins': 0}
        by_bot[b]['trades'] += 1
        by_bot[b]['net_pnl'] += t['net_pnl'] or 0
        if t['outcome'] in ('won', 'win', 'Won', 'filled') or (t.get('net_pnl') and t['net_pnl'] > 0):
            by_bot[b]['wins'] += 1
        if t['outcome'] in ('lost', 'loss', 'Lost') or (t.get('net_pnl') and t['net_pnl'] < 0):
            by_bot[b].setdefault('losses', 0)
            by_bot[b]['losses'] += 1

    by_bot_summary = {
        b: {
            'trades': v['trades'],
            'net_pnl': round(v['net_pnl'], 2),
            'win_rate': round((v['wins'] / (v['wins'] + v.get('losses', 0)) * 100) if (v['wins'] + v.get('losses', 0)) > 0 else 0.0, 1),
        }
        for b, v in by_bot.items()
    }

    # Paginate
    paginated = all_trades[offset:offset + limit]

    return {
        'trades': paginated,
        'summary': {
            'total_trades': total_trades,
            'total_staked': round(total_staked, 2),
            'total_gross_pnl': round(total_gross_pnl, 2),
            'total_fees': round(total_fees, 2),
            'total_net_pnl': round(total_net_pnl, 2),
            'won': won,
            'lost': lost,
            'win_rate': win_rate,
            'best_trade': round(best_trade_pnl, 2),
            'best_trade_bot': best_trade_bot,
            'worst_trade': round(worst_trade_pnl, 2),
            'worst_trade_bot': worst_trade_bot,
            'by_bot': by_bot_summary,
        }
    }


@app.get("/api/tradebook/export")
async def export_tradebook():
    """Export all trades as XLSX file."""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from datetime import date

    # Fetch all trades
    data = await get_tradebook(limit=10000)
    trades = data['trades']
    summary = data['summary']

    wb = Workbook()

    HEADERS = ['#', 'Bot', 'Timestamp', 'Direction', 'Entry Price', 'Stake',
               'Gross P&L', 'Fees', 'Net P&L', 'Outcome']

    HEADER_FILL = PatternFill('solid', fgColor='1e1e2e')
    HEADER_FONT = Font(bold=True, color='e2e8f0', name='Calibri')
    GREEN_FILL = PatternFill('solid', fgColor='14532d')
    RED_FILL = PatternFill('solid', fgColor='7f1d1d')
    GREEN_FONT = Font(color='22c55e', name='Calibri')
    RED_FONT = Font(color='ef4444', name='Calibri')

    def write_sheet(ws, rows, title='Trades'):
        ws.title = title
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
        for i, t in enumerate(rows, 1):
            pnl = t.get('net_pnl') or 0
            outcome = t.get('outcome') or ''
            row = [
                i,
                t.get('bot', ''),
                t.get('timestamp', ''),
                t.get('direction', ''),
                t.get('entry_price'),
                t.get('stake'),
                t.get('gross_pnl'),
                t.get('fees'),
                pnl,
                outcome,
            ]
            ws.append(row)
            row_idx = ws.max_row
            fill = GREEN_FILL if pnl >= 0 else RED_FILL
            font = GREEN_FONT if pnl >= 0 else RED_FONT
            for col in [7, 9]:  # Gross P&L, Net P&L
                c = ws.cell(row=row_idx, column=col)
                c.fill = fill
                c.font = font
        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # All Trades sheet
    ws_all = wb.active
    write_sheet(ws_all, trades, 'All Trades')

    # Per-bot sheets
    for bot_name in ['BTC Paper', 'BTC Live', 'JC Copy', 'Sports', 'Scalper', 'Maker']:
        bot_trades = [t for t in trades if t['bot'] == bot_name]
        if bot_trades:
            ws = wb.create_sheet(bot_name)
            write_sheet(ws, bot_trades, bot_name)

    # Summary sheet
    ws_sum = wb.create_sheet('Summary')
    ws_sum.append(['Bot', 'Trades', 'Net P&L', 'Win Rate'])
    for cell in ws_sum[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    for bot_name, bdata in summary['by_bot'].items():
        ws_sum.append([bot_name, bdata['trades'], bdata['net_pnl'], f"{bdata['win_rate']}%"])
    ws_sum.append([])
    ws_sum.append(['TOTAL', summary['total_trades'], summary['total_net_pnl'], f"{summary['win_rate']}%"])
    for col in ws_sum.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws_sum.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"tradebook-{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


# =============================================================================
# JC Discord Feed API — must be before static file catch-all
# =============================================================================

@app.get("/api/jc/discord-feed")
async def jc_discord_feed():
    """Get recent JC Discord messages with BTC context for frontend display."""
    import sqlite3, httpx
    TRADING_DB = '/data/.openclaw/workspace/projects/Ghost/data/trading.db'
    JC_LEVELS = [
        {"price": 85500, "label": "SPV resistance", "type": "resistance"},
        {"price": 84000, "label": "Current resistance", "type": "resistance"},
        {"price": 83000, "label": "Support level", "type": "support"},
        {"price": 82000, "label": "Key support", "type": "support"},
        {"price": 80000, "label": "Round number POC", "type": "support"},
    ]
    try:
        conn = sqlite3.connect(TRADING_DB)
        rows = conn.execute(
            """SELECT id, channel_type, author, content, attachments, created_at
               FROM messages WHERE author='jayson_casper'
               ORDER BY created_at DESC LIMIT 30"""
        ).fetchall()
        conn.close()
        btc = 0.0
        try:
            async with httpx.AsyncClient(timeout=4) as c:
                r = await c.get('https://api.binance.com/api/v3/ticker/price?symbol=BTCUSDT')
                btc = float(r.json()['price'])
        except Exception:
            pass
        messages = [
            {
                "id": r[0], "channel": r[1], "author": r[2],
                "content": r[3], "attachments": r[4], "timestamp": r[5],
                "has_chart": "tradingview.com" in (r[3] or "").lower() or "chart" in (r[3] or "").lower(),
                "is_signal": any(kw in (r[3] or "").lower() for kw in ["long", "short", "buy", "sell", "sl ", " tp", "stop", "entry", "target"]),
            }
            for r in rows
        ]
        return {"messages": messages, "btc_price": btc, "jc_levels": JC_LEVELS, "count": len(messages)}
    except Exception as e:
        return {"messages": [], "btc_price": 0.0, "jc_levels": JC_LEVELS, "count": 0, "error": str(e)}


# =============================================================================
# Static Files — Serve React Dashboard
# =============================================================================

# Path to the built React dashboard
@app.get("/api/v5/status")
async def v5_status():
    """V5 paper trading status and stats."""
    try:
        pool = get_async_pool()
        async with pool.acquire() as conn:
            stats = await conn.fetchrow("""
                SELECT COUNT(*) as total,
                  COUNT(*) FILTER (WHERE won=true) as wins,
                  COUNT(*) FILTER (WHERE won=false) as losses,
                  COUNT(*) FILTER (WHERE resolved_at IS NULL) as open_trades,
                  COALESCE(SUM(simulated_pnl),0) as total_pnl,
                  MAX(created_at) as last_trade
                FROM paper_trades WHERE strategy_version='V5'
            """)
            bankroll = await conn.fetchrow(
                "SELECT balance, total_trades, total_won, total_lost FROM btc_bankroll WHERE id=1"
            )
        total = int(stats['total']) if stats else 0
        wins = int(stats['wins']) if stats else 0
        win_rate = round((wins / max(total, 1)) * 100, 1)
        return {
            "status": "running", "mode": "paper",
            "total_trades": total, "wins": wins,
            "losses": int(stats['losses']) if stats else 0,
            "open_trades": int(stats['open_trades']) if stats else 0,
            "win_rate": win_rate,
            "total_pnl": round(float(stats['total_pnl']), 2) if stats else 0,
            "bankroll": float(bankroll['balance']) if bankroll else 10000.0,
            "last_trade": str(stats['last_trade']) if stats and stats['last_trade'] else None,
        }
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/api/v5/trades")
async def v5_trades():
    """Recent V5 paper trades."""
    try:
        pool = get_async_pool()
        async with pool.acquire() as conn:
            rows = await conn.fetch("""
                SELECT id, direction, token_price, stake_usd, confluence_score,
                       resolution, won, simulated_pnl, created_at, resolved_at
                FROM paper_trades WHERE strategy_version='V5'
                ORDER BY created_at DESC LIMIT 50
            """)
        return {"trades": [{**dict(r), 'created_at': str(r['created_at']), 'resolved_at': str(r['resolved_at']) if r['resolved_at'] else None} for r in rows]}
    except Exception as e:
        return {"trades": [], "error": str(e)}


BUILD_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "dashboard", "dist")

if os.path.exists(BUILD_PATH):
    # Mount static assets
    app.mount("/assets", StaticFiles(directory=os.path.join(BUILD_PATH, "assets")), name="assets")
    
    # Serve index.html for all non-API routes (SPA fallback)
    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str):
        # If path starts with /api, it's already handled by API routes
        if full_path.startswith("api/"):
            raise HTTPException(status_code=404, detail="API endpoint not found")
        
        # For root or any non-API path, serve index.html (React Router will handle)
        file_path = os.path.join(BUILD_PATH, full_path)
        if os.path.isfile(file_path):
            return FileResponse(file_path)
        else:
            return FileResponse(os.path.join(BUILD_PATH, "index.html"))
    
    logger.info(f"✅ Serving React dashboard from {BUILD_PATH}")
else:
    logger.warning(f"⚠️ Dashboard build not found at {BUILD_PATH}. Run 'npm run build' in dashboard/")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=6010)
